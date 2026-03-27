from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Query, Body
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt
import csv
import io
from zoneinfo import ZoneInfo
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import ssl
import openpyxl
from openpyxl.utils import get_column_letter

# IST Timezone
IST = ZoneInfo("Asia/Kolkata")

def get_ist_now():
    """Get current datetime in IST timezone"""
    return datetime.now(IST)

def get_ist_now_iso():
    """Get current datetime in IST as ISO string"""
    return datetime.now(IST).isoformat()

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
SECRET_KEY = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480

security = HTTPBearer()

app = FastAPI()
api_router = APIRouter(prefix="/api")

# Pydantic Models
class UserLogin(BaseModel):
    username: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: dict

class DepartmentCreate(BaseModel):
    name: str
    description: Optional[str] = ""

class Department(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    description: Optional[str] = ""
    is_active: bool = True
    created_at: str
    updated_at: str

class EmployeeCreate(BaseModel):
    employee_id: str
    name: str
    email: EmailStr
    phone: str
    department_ids: List[str]  # Changed to support multiple departments
    experience: str
    password: str
    joining_date: str
    birth_date: Optional[str] = None  # New field
    team_leader_ids: Optional[List[str]] = []  # Team leaders (fixed per employee)
    bank_id: Optional[str] = None  # Bank selection
    pt: Optional[str] = None  # PT field
    esic: Optional[str] = None  # ESIC field
    epf: Optional[str] = None  # EPF field
    cpf: Optional[str] = None  # CPF field
    salary: Optional[str] = None  # Salary field

class EmployeeUpdate(BaseModel):
    employee_id: Optional[str] = None  # Now editable
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    department_ids: Optional[List[str]] = None  # Changed to support multiple departments
    experience: Optional[str] = None
    password: Optional[str] = None
    birth_date: Optional[str] = None  # New field
    joining_date: Optional[str] = None  # Now editable
    team_leader_ids: Optional[List[str]] = None  # Team leaders (fixed per employee)
    bank_id: Optional[str] = None  # Bank selection
    bank_account_number: Optional[str] = None  # Bank account number
    ifsc_code: Optional[str] = None  # IFSC code
    pt: Optional[str] = None  # PT field
    esic: Optional[str] = None  # ESIC field
    epf: Optional[str] = None  # EPF field
    cpf: Optional[str] = None  # CPF field
    salary: Optional[str] = None  # Salary field

class Employee(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    name: str
    email: str
    phone: str
    department_ids: List[str] = []  # Changed to support multiple departments
    department_id: Optional[str] = ""  # Keep for backward compatibility
    role: str
    experience: Optional[str] = ""
    joining_date: str
    birth_date: Optional[str] = None  # New field
    status: str
    probation_end_date: str
    annual_pl_allocation: int
    pl_taken: float
    cl_taken: float
    created_at: str
    updated_at: str
    team_leader_ids: List[str] = []  # Assigned team leaders
    bank_id: Optional[str] = None  # Bank selection
    bank_account_number: Optional[str] = None  # Bank account number
    ifsc_code: Optional[str] = None  # IFSC code
    pt: Optional[str] = None  # PT field
    esic: Optional[str] = None  # ESIC field
    epf: Optional[str] = None  # EPF field
    cpf: Optional[str] = None  # CPF field
    salary: Optional[str] = None  # Salary field
    plain_password: Optional[str] = None  # Plain text password for admin visibility

class ProjectCreate(BaseModel):
    name: str
    type: str
    project_code: str
    start_date: str
    end_date: Optional[str] = ""  # Made optional - not required for late logic
    completed_hours: float = 0.0
    assigned_employees: List[str]
    status: str
    client_username: str
    scope_of_work: str
    timesheet_link: str
    is_late: Optional[bool] = False  # Manual late marking

class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    type: Optional[str] = None
    project_code: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    completed_hours: Optional[float] = None
    assigned_employees: Optional[List[str]] = None
    status: Optional[str] = None
    client_username: Optional[str] = None
    scope_of_work: Optional[str] = None
    timesheet_link: Optional[str] = None
    is_late: Optional[bool] = None  # Manual late marking

class Project(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    type: str
    project_code: str
    start_date: str
    end_date: Optional[str] = ""
    completed_hours: float
    assigned_employees: List[str]
    status: str
    client_username: str
    scope_of_work: str
    timesheet_link: str
    is_late: bool = False
    created_at: str
    updated_at: str

class WorkEntryCreate(BaseModel):
    project_code: str  # Changed from project_id to project_code
    hours: float
    work_details: str
    date: str

# Holiday Models
class HolidayCreate(BaseModel):
    name: str
    date: str
    description: Optional[str] = ""

class Holiday(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    date: str
    description: Optional[str] = ""
    created_at: str

class WorkEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    date: str
    project_code: str  # Changed from project_id to project_code
    hours: float
    work_details: str
    created_at: str

# Weekend/Holiday Approval Models
class WeekendApprovalCreate(BaseModel):
    project_code: str
    hours: float
    work_details: str
    date: str

class WeekendApprovalAction(BaseModel):
    approved_date: Optional[str] = None  # Admin can edit date
    approved_hours: Optional[float] = None  # Admin can edit hours
    rejection_reason: Optional[str] = ""
    is_compensation: Optional[bool] = False  # For future use

class LeaveDateInput(BaseModel):
    date: str
    day_type: str = "full"  # full or half

class LeaveApplication(BaseModel):
    from_date: str
    to_date: str
    leave_type: Optional[str] = None  # Optional - Admin/HR will set this during approval
    reason: str
    leave_dates: Optional[List[LeaveDateInput]] = None  # New: support multiple dates with half-day

class LeaveDateType(BaseModel):
    date: str
    leave_type: str  # PL, CL, Half PL, Half CL, PL/2 & CL/2, or "Rejected"
    reject_reason: Optional[str] = ""  # Mandatory if leave_type is "Rejected"

class LeaveApproval(BaseModel):
    status: str  # approved, rejected, or partial (for partial approval/rejection)
    comments: Optional[str] = ""
    reject_reason: Optional[str] = ""
    leave_dates: Optional[List[LeaveDateType]] = []

class AdminUserCreate(BaseModel):
    username: str
    email: str
    password: str
    role: str = "admin"  # admin or hr

# Email Configuration Model
class EmailConfigUpdate(BaseModel):
    smtp_host: str = "smtp.gmail.com"
    smtp_port: int = 587
    smtp_email: str = "hr.zestbrains@gmail.com"
    smtp_password: str = ""
    enable_ssl: bool = True
    cc_emails: str = ""  # Comma-separated emails
    is_enabled: bool = False

# Email Sending Utility Function
async def send_leave_notification_email(
    employee_email: str,
    employee_name: str,
    leave_type: str,
    applied_dates: List[str],
    approved_dates: List[dict],  # [{date, type}]
    rejected_dates: List[dict],  # [{date, reason}]
    status: str,  # approved, rejected, partial
    approved_by: str,
    comments: str = ""
):
    """Send leave approval/rejection email to employee"""
    try:
        # Get email configuration from database
        email_config = await db.email_config.find_one({}, {"_id": 0})
        if not email_config or not email_config.get("is_enabled"):
            logging.info("Email notifications disabled or not configured")
            return False
        
        smtp_host = email_config.get("smtp_host", "smtp.gmail.com")
        smtp_port = email_config.get("smtp_port", 587)
        smtp_email = email_config.get("smtp_email", "hr.zestbrains@gmail.com")
        smtp_password = email_config.get("smtp_password", "")
        enable_ssl = email_config.get("enable_ssl", True)
        cc_emails = email_config.get("cc_emails", "")
        
        if not smtp_password:
            logging.warning("SMTP password not configured")
            return False
        
        # Build email content
        now_ist = get_ist_now().strftime("%d %b %Y, %I:%M %p IST")
        
        subject = f"Leave Application {status.upper()} - {employee_name}"
        
        # Build HTML email body
        approved_dates_html = ""
        if approved_dates:
            approved_dates_html = "<h3 style='color: #16a34a;'>✓ Approved Dates:</h3><ul>"
            for d in approved_dates:
                approved_dates_html += f"<li>{d['date']} - {d['type']}</li>"
            approved_dates_html += "</ul>"
        
        rejected_dates_html = ""
        if rejected_dates:
            rejected_dates_html = "<h3 style='color: #dc2626;'>✗ Rejected Dates:</h3><ul>"
            for d in rejected_dates:
                rejected_dates_html += f"<li>{d['date']} - Reason: {d['reason']}</li>"
            rejected_dates_html += "</ul>"
        
        status_color = "#16a34a" if status == "approved" else "#dc2626" if status == "rejected" else "#d97706"
        status_text = "APPROVED" if status == "approved" else "REJECTED" if status == "rejected" else "PARTIALLY APPROVED"
        
        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <h2 style="color: #4f46e5;">Zestbrains - Leave Application Update</h2>
                <hr style="border: 1px solid #e5e7eb;">
                
                <p>Dear <strong>{employee_name}</strong>,</p>
                
                <p>Your leave application has been <span style="color: {status_color}; font-weight: bold;">{status_text}</span>.</p>
                
                <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
                    <tr>
                        <td style="padding: 8px; border: 1px solid #e5e7eb; background: #f9fafb;"><strong>Employee Name</strong></td>
                        <td style="padding: 8px; border: 1px solid #e5e7eb;">{employee_name}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #e5e7eb; background: #f9fafb;"><strong>Leave Type</strong></td>
                        <td style="padding: 8px; border: 1px solid #e5e7eb;">{leave_type or 'As per dates below'}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #e5e7eb; background: #f9fafb;"><strong>Applied Dates</strong></td>
                        <td style="padding: 8px; border: 1px solid #e5e7eb;">{', '.join(applied_dates)}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #e5e7eb; background: #f9fafb;"><strong>Approved By</strong></td>
                        <td style="padding: 8px; border: 1px solid #e5e7eb;">{approved_by}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #e5e7eb; background: #f9fafb;"><strong>Date of Action</strong></td>
                        <td style="padding: 8px; border: 1px solid #e5e7eb;">{now_ist}</td>
                    </tr>
                </table>
                
                {approved_dates_html}
                {rejected_dates_html}
                
                {"<p><strong>Comments:</strong> " + comments + "</p>" if comments else ""}
                
                <hr style="border: 1px solid #e5e7eb; margin-top: 30px;">
                <p style="color: #6b7280; font-size: 12px;">
                    This is an automated email from Zestbrains HR Portal. Please do not reply to this email.
                </p>
            </div>
        </body>
        </html>
        """
        
        # Create message
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = smtp_email
        msg['To'] = employee_email
        
        # Add CC if configured
        cc_list = [e.strip() for e in cc_emails.split(',') if e.strip()]
        if cc_list:
            msg['Cc'] = ', '.join(cc_list)
        
        msg.attach(MIMEText(html_body, 'html'))
        
        # Send email
        all_recipients = [employee_email] + cc_list
        
        if enable_ssl:
            context = ssl.create_default_context()
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls(context=context)
                server.login(smtp_email, smtp_password)
                server.sendmail(smtp_email, all_recipients, msg.as_string())
        else:
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.login(smtp_email, smtp_password)
                server.sendmail(smtp_email, all_recipients, msg.as_string())
        
        logging.info(f"Leave notification email sent to {employee_email}")
        return True
        
    except Exception as e:
        logging.error(f"Failed to send email: {str(e)}")
        return False

# Utility Functions
def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = get_ist_now() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        if username is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user = await db.users.find_one({"username": username}, {"_id": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except (jwt.PyJWTError, jwt.DecodeError, jwt.ExpiredSignatureError, Exception):
        raise HTTPException(status_code=401, detail="Invalid token")

def require_role(allowed_roles: List[str]):
    async def role_checker(user: dict = Depends(get_current_user)):
        if user["role"] not in allowed_roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return role_checker

# Authentication Routes
@api_router.post("/auth/login", response_model=Token)
async def login(credentials: UserLogin):
    # Try to find user by username first (for admin/hr)
    user = await db.users.find_one({"username": credentials.username}, {"_id": 0})
    
    # If not found by username, try email (for employees)
    if not user:
        user = await db.users.find_one({"email": credentials.username}, {"_id": 0})
    
    if not user or not verify_password(credentials.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not user.get("is_active", True):
        raise HTTPException(status_code=401, detail="Account is inactive")
    
    access_token = create_access_token({"sub": user["username"], "role": user["role"]})
    user_data = {k: v for k, v in user.items() if k != "password_hash"}
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": user_data
    }

@api_router.get("/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    return {k: v for k, v in user.items() if k != "password_hash"}

# Admin User Management Routes
@api_router.get("/admin/users")
async def get_admin_users(user: dict = Depends(require_role(["admin"]))):
    """Get all admin and HR users"""
    users = await db.users.find(
        {"role": {"$in": ["admin", "hr"]}},
        {"_id": 0, "password_hash": 0}
    ).to_list(1000)
    return users

@api_router.post("/admin/users")
async def create_admin_user(admin_user: AdminUserCreate, user: dict = Depends(require_role(["admin"]))):
    """Create a new admin or HR user"""
    # Check if username already exists
    existing = await db.users.find_one({"username": admin_user.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Check if email already exists
    existing_email = await db.users.find_one({"email": admin_user.email})
    if existing_email:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    # Validate role
    if admin_user.role not in ["admin", "hr"]:
        raise HTTPException(status_code=400, detail="Role must be 'admin' or 'hr'")
    
    now = get_ist_now_iso()
    user_id = str(uuid.uuid4())
    
    user_doc = {
        "id": user_id,
        "username": admin_user.username,
        "email": admin_user.email,
        "password_hash": hash_password(admin_user.password),
        "role": admin_user.role,
        "employee_id": f"{admin_user.role.upper()}{str(uuid.uuid4())[:4].upper()}",
        "is_active": True,
        "created_at": now,
        "updated_at": now
    }
    
    await db.users.insert_one(user_doc)
    
    return {"id": user_id, "username": admin_user.username, "email": admin_user.email, "role": admin_user.role}

@api_router.delete("/admin/users/{user_id}")
async def delete_admin_user(user_id: str, user: dict = Depends(require_role(["admin"]))):
    """Delete an admin or HR user"""
    # Don't allow deleting the current logged-in user
    target_user = await db.users.find_one({"id": user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if target_user["username"] == user["username"]:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    await db.users.delete_one({"id": user_id})
    return {"message": "User deleted successfully"}

# Department Routes
@api_router.get("/departments", response_model=List[Department])
async def get_departments(user: dict = Depends(get_current_user)):
    departments = await db.departments.find({}, {"_id": 0}).to_list(1000)
    return departments

@api_router.post("/departments", response_model=Department)
async def create_department(dept: DepartmentCreate, user: dict = Depends(require_role(["admin"]))):
    dept_id = str(uuid.uuid4())
    now = get_ist_now_iso()
    
    dept_doc = {
        "id": dept_id,
        "name": dept.name,
        "description": dept.description,
        "is_active": True,
        "created_at": now,
        "updated_at": now
    }
    
    await db.departments.insert_one(dept_doc)
    created_dept = await db.departments.find_one({"id": dept_id}, {"_id": 0})
    return created_dept

@api_router.put("/departments/{dept_id}", response_model=Department)
async def update_department(dept_id: str, dept: DepartmentCreate, user: dict = Depends(require_role(["admin"]))):
    now = get_ist_now_iso()
    update_data = {
        "name": dept.name,
        "description": dept.description,
        "updated_at": now
    }
    
    result = await db.departments.update_one({"id": dept_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Department not found")
    
    updated = await db.departments.find_one({"id": dept_id}, {"_id": 0})
    return updated

@api_router.put("/departments/{dept_id}/status")
async def update_department_status(dept_id: str, is_active: bool, user: dict = Depends(require_role(["admin"]))):
    result = await db.departments.update_one(
        {"id": dept_id},
        {"$set": {"is_active": is_active, "updated_at": get_ist_now_iso()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Department not found")
    return {"message": "Status updated"}

@api_router.delete("/departments/{dept_id}")
async def delete_department(dept_id: str, user: dict = Depends(require_role(["admin"]))):
    result = await db.departments.delete_one({"id": dept_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Department not found")
    return {"message": "Department deleted"}

# Employee Routes
@api_router.get("/employees", response_model=List[Employee])
async def get_employees(status: Optional[str] = None, user: dict = Depends(get_current_user)):
    query = {}
    if status and status != "all":
        query["status"] = status
    
    employees = await db.employees.find(query, {"_id": 0}).to_list(1000)
    return employees

@api_router.get("/employees/{emp_id}", response_model=Employee)
async def get_employee(emp_id: str, user: dict = Depends(get_current_user)):
    employee = await db.employees.find_one({"id": emp_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    return employee

@api_router.post("/employees", response_model=Employee)
async def create_employee(emp: EmployeeCreate, user: dict = Depends(require_role(["admin"]))):
    # Check if employee_id already exists
    existing = await db.employees.find_one({"employee_id": emp.employee_id})
    if existing:
        raise HTTPException(status_code=400, detail="Employee ID already exists")
    
    # Auto-generate username from email
    username = emp.email.split('@')[0]
    
    # Check if username exists, add number if needed
    base_username = username
    counter = 1
    while await db.users.find_one({"username": username}):
        username = f"{base_username}{counter}"
        counter += 1
    
    emp_id = str(uuid.uuid4())
    now = get_ist_now_iso()
    joining_date = datetime.fromisoformat(emp.joining_date)
    probation_end = (joining_date + timedelta(days=90)).isoformat()
    
    # Create user account with employee role by default
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "username": username,
        "email": emp.email,
        "password_hash": hash_password(emp.password),
        "role": "employee",
        "employee_id": emp.employee_id,
        "is_active": True,
        "created_at": now
    }
    await db.users.insert_one(user_doc)
    
    emp_doc = {
        "id": emp_id,
        "employee_id": emp.employee_id,
        "name": emp.name,
        "email": emp.email,
        "phone": emp.phone,
        "department_ids": emp.department_ids,  # Multiple departments
        "department_id": emp.department_ids[0] if emp.department_ids else "",  # Backward compatibility
        "role": "employee",
        "experience": emp.experience,
        "joining_date": emp.joining_date,
        "birth_date": emp.birth_date,  # Birth date
        "team_leader_ids": emp.team_leader_ids or [],  # Assigned team leaders
        "status": "active",
        "probation_end_date": probation_end,
        "annual_pl_allocation": 16,
        "pl_taken": 0.0,
        "cl_taken": 0.0,
        "bank_id": emp.bank_id,  # Bank ID
        "pt": emp.pt,  # PT
        "esic": emp.esic,  # ESIC
        "epf": emp.epf,  # EPF
        "cpf": emp.cpf,  # CPF
        "salary": emp.salary,  # Salary
        "plain_password": emp.password,  # Store plain text password for admin visibility
        "created_at": now,
        "updated_at": now
    }
    
    await db.employees.insert_one(emp_doc)
    created_emp = await db.employees.find_one({"id": emp_id}, {"_id": 0})
    return created_emp

@api_router.put("/employees/{emp_id}", response_model=Employee)
async def update_employee(emp_id: str, emp: EmployeeUpdate, user: dict = Depends(require_role(["admin"]))):
    now = get_ist_now_iso()
    update_data = {"updated_at": now}
    
    employee = await db.employees.find_one({"id": emp_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    old_employee_id = employee["employee_id"]
    
    if emp.employee_id:
        # Check if new employee_id already exists
        if emp.employee_id != old_employee_id:
            existing = await db.employees.find_one({"employee_id": emp.employee_id, "id": {"$ne": emp_id}})
            if existing:
                raise HTTPException(status_code=400, detail="Employee ID already exists")
        update_data["employee_id"] = emp.employee_id
    if emp.name:
        update_data["name"] = emp.name
    if emp.email:
        update_data["email"] = emp.email
    if emp.phone:
        update_data["phone"] = emp.phone
    if emp.department_ids is not None:
        update_data["department_ids"] = emp.department_ids
        update_data["department_id"] = emp.department_ids[0] if emp.department_ids else ""
    if emp.experience:
        update_data["experience"] = emp.experience
    if emp.birth_date is not None:
        update_data["birth_date"] = emp.birth_date
    if emp.joining_date:
        update_data["joining_date"] = emp.joining_date
    if emp.team_leader_ids is not None:
        update_data["team_leader_ids"] = emp.team_leader_ids
    if emp.bank_id is not None:
        update_data["bank_id"] = emp.bank_id
    if emp.pt is not None:
        update_data["pt"] = emp.pt
    if emp.esic is not None:
        update_data["esic"] = emp.esic
    if emp.epf is not None:
        update_data["epf"] = emp.epf
    if emp.cpf is not None:
        update_data["cpf"] = emp.cpf
    if emp.salary is not None:
        update_data["salary"] = emp.salary
    
    result = await db.employees.update_one({"id": emp_id}, {"$set": update_data})
    
    # Update user record if employee_id changed
    if emp.employee_id and emp.employee_id != old_employee_id:
        await db.users.update_one(
            {"employee_id": old_employee_id},
            {"$set": {"employee_id": emp.employee_id}}
        )
        # Also update work_entries and leave_applications
        await db.work_entries.update_many(
            {"employee_id": old_employee_id},
            {"$set": {"employee_id": emp.employee_id}}
        )
        await db.leave_applications.update_many(
            {"employee_id": old_employee_id},
            {"$set": {"employee_id": emp.employee_id}}
        )
        await db.leave_records.update_many(
            {"employee_id": old_employee_id},
            {"$set": {"employee_id": emp.employee_id}}
        )
    
    # Update user if password changed and also store plain password
    if emp.password:
        current_emp_id = emp.employee_id if emp.employee_id else old_employee_id
        await db.users.update_one(
            {"employee_id": current_emp_id},
            {"$set": {"password_hash": hash_password(emp.password)}}
        )
        # Also update plain_password in employees collection
        await db.employees.update_one(
            {"id": emp_id},
            {"$set": {"plain_password": emp.password}}
        )
    
    updated = await db.employees.find_one({"id": emp_id}, {"_id": 0})
    return updated

@api_router.put("/employees/{emp_id}/status")
async def update_employee_status(emp_id: str, status: str, user: dict = Depends(require_role(["admin", "hr"]))):
    # Allow toggling between active and ex-employee
    if status not in ["active", "inactive", "ex-employee"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    result = await db.employees.update_one(
        {"id": emp_id},
        {"$set": {"status": status, "updated_at": get_ist_now_iso()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Update user active status
    employee = await db.employees.find_one({"id": emp_id})
    is_active = status == "active"
    await db.users.update_one(
        {"employee_id": employee["employee_id"]},
        {"$set": {"is_active": is_active}}
    )
    
    return {"message": "Status updated"}

@api_router.delete("/employees/{emp_id}")
async def delete_employee(emp_id: str, user: dict = Depends(require_role(["admin"]))):
    employee = await db.employees.find_one({"id": emp_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Delete user account
    await db.users.delete_one({"employee_id": employee["employee_id"]})
    
    result = await db.employees.delete_one({"id": emp_id})
    return {"message": "Employee deleted"}

# Project Routes
@api_router.get("/projects", response_model=List[Project])
async def get_projects(user: dict = Depends(get_current_user)):
    if user["role"] == "employee":
        # Employees only see assigned projects
        employee = await db.employees.find_one({"email": user["email"]})
        projects = await db.projects.find(
            {"assigned_employees": employee["employee_id"]},
            {"_id": 0}
        ).to_list(1000)
    else:
        projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    
    # Custom sort: zb_new_* first (descending), then zb_old_* (descending)
    def sort_key(project):
        code = project.get("project_code", "")
        # Extract numeric part from project code (e.g., "zb_new_100" -> 100)
        try:
            num = int(code.split("_")[-1])
        except:
            num = 0
        
        # "new" should come before "old", so we use 0 for new, 1 for old
        if "new" in code.lower():
            return (0, -num)  # 0 = first priority, -num for descending
        elif "old" in code.lower():
            return (1, -num)  # 1 = second priority, -num for descending
        else:
            return (2, -num)  # Other codes at the end
    
    projects.sort(key=sort_key)
    return projects

@api_router.get("/projects/pm-view")
async def get_pm_projects_view(user: dict = Depends(get_current_user)):
    """Get projects assigned to PM employee with work hours summary"""
    if user["role"] != "employee":
        raise HTTPException(status_code=403, detail="This endpoint is for employees only")
    
    # Get employee details
    employee = await db.employees.find_one({"email": user["email"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Check if employee is in PM department
    pm_dept = await db.departments.find_one({"name": {"$regex": "^PM$", "$options": "i"}})
    if not pm_dept:
        raise HTTPException(status_code=404, detail="PM department not found")
    
    pm_id = pm_dept.get("id")
    emp_depts = employee.get("department_ids", [])
    
    if pm_id not in emp_depts:
        raise HTTPException(status_code=403, detail="You are not in the PM department")
    
    emp_id = employee.get("employee_id")
    
    # Get all projects assigned to this employee
    projects = await db.projects.find(
        {"assigned_employees": emp_id},
        {"_id": 0}
    ).to_list(1000)
    
    if not projects:
        return []
    
    # Get all work entries for these projects
    project_codes = [p["project_code"] for p in projects]
    work_entries = await db.work_entries.find(
        {"project_code": {"$in": project_codes}},
        {"_id": 0}
    ).to_list(100000)
    
    # Get all employees for name mapping
    all_employees = await db.employees.find({}, {"_id": 0, "employee_id": 1, "name": 1}).to_list(1000)
    emp_map = {e["employee_id"]: e["name"] for e in all_employees}
    
    # Group work entries by project
    project_hours = {}
    for entry in work_entries:
        pc = entry["project_code"]
        if pc not in project_hours:
            project_hours[pc] = {"total": 0, "developers": {}}
        project_hours[pc]["total"] += entry["hours"]
        
        dev_id = entry["employee_id"]
        if dev_id not in project_hours[pc]["developers"]:
            project_hours[pc]["developers"][dev_id] = {
                "name": emp_map.get(dev_id, f"Emp #{dev_id}"),
                "hours": 0
            }
        project_hours[pc]["developers"][dev_id]["hours"] += entry["hours"]
    
    # Enrich projects with hours data
    result = []
    for proj in projects:
        pc = proj["project_code"]
        hours_data = project_hours.get(pc, {"total": 0, "developers": {}})
        
        # Sort developers by hours descending
        devs = sorted(
            hours_data["developers"].values(),
            key=lambda x: x["hours"],
            reverse=True
        )
        
        # Use completed_hours from project (same as admin view) for total display
        # But keep developer breakdown from work_entries for detailed view
        result.append({
            **proj,
            "total_hours": proj.get("completed_hours", 0),  # Use project's completed_hours like admin
            "developer_count": len(devs),
            "developers": devs
        })
    
    # Sort projects: zb_new_* first (descending), then zb_old_* (descending)
    def pm_sort_key(project):
        code = project.get("project_code", "")
        try:
            num = int(code.split("_")[-1])
        except:
            num = 0
        if "new" in code.lower():
            return (0, -num)
        elif "old" in code.lower():
            return (1, -num)
        else:
            return (2, -num)
    
    result.sort(key=pm_sort_key)
    return result

@api_router.get("/projects/team-view")
async def get_team_projects_view(user: dict = Depends(get_current_user)):
    """Get projects for team leader - includes own projects and team members' projects with work entries"""
    if user["role"] != "employee":
        raise HTTPException(status_code=403, detail="This endpoint is for employees only")
    
    # Get employee details
    employee = await db.employees.find_one({"email": user["email"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    emp_id = employee.get("employee_id")
    
    # Find all team members (employees who have this employee as team leader)
    team_members = await db.employees.find(
        {"team_leader_ids": emp_id},
        {"_id": 0}
    ).to_list(1000)
    
    if not team_members:
        raise HTTPException(status_code=403, detail="You are not a team leader for any employees")
    
    # Get team member employee IDs
    team_member_ids = [m["employee_id"] for m in team_members]
    
    # Get all projects assigned to team members ONLY (not team leader's own projects)
    projects = await db.projects.find(
        {"assigned_employees": {"$in": team_member_ids}},
        {"_id": 0}
    ).to_list(1000)
    
    if not projects:
        return []
    
    # Get all work entries for these projects
    project_codes = [p["project_code"] for p in projects]
    work_entries = await db.work_entries.find(
        {"project_code": {"$in": project_codes}},
        {"_id": 0}
    ).to_list(100000)
    
    # Get all employees for name mapping
    all_employees = await db.employees.find({}, {"_id": 0, "employee_id": 1, "name": 1}).to_list(1000)
    emp_map = {e["employee_id"]: e["name"] for e in all_employees}
    
    # Group work entries by project
    project_hours = {}
    for entry in work_entries:
        pc = entry["project_code"]
        if pc not in project_hours:
            project_hours[pc] = {"total": 0, "developers": {}}
        project_hours[pc]["total"] += entry["hours"]
        
        dev_id = entry["employee_id"]
        if dev_id not in project_hours[pc]["developers"]:
            project_hours[pc]["developers"][dev_id] = {
                "employee_id": dev_id,
                "name": emp_map.get(dev_id, f"Emp #{dev_id}"),
                "hours": 0,
                "is_team_member": dev_id in team_member_ids,
                "is_self": dev_id == emp_id
            }
        project_hours[pc]["developers"][dev_id]["hours"] += entry["hours"]
    
    # Enrich projects with hours data
    result = []
    for proj in projects:
        pc = proj["project_code"]
        hours_data = project_hours.get(pc, {"total": 0, "developers": {}})
        
        # Sort developers by hours descending
        devs = sorted(
            hours_data["developers"].values(),
            key=lambda x: x["hours"],
            reverse=True
        )
        
        result.append({
            **proj,
            "total_hours": proj.get("completed_hours", 0),
            "developer_count": len(devs),
            "developers": devs,
            "team_member_count": len([d for d in devs if d["is_team_member"]])
        })
    
    # Sort projects: zb_new_* first (descending), then zb_old_* (descending)
    def tl_sort_key(project):
        code = project.get("project_code", "")
        try:
            num = int(code.split("_")[-1])
        except:
            num = 0
        if "new" in code.lower():
            return (0, -num)
        elif "old" in code.lower():
            return (1, -num)
        else:
            return (2, -num)
    
    result.sort(key=tl_sort_key)
    return result

@api_router.get("/projects/{proj_id}", response_model=Project)
async def get_project(proj_id: str, user: dict = Depends(get_current_user)):
    project = await db.projects.find_one({"id": proj_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project

@api_router.post("/projects", response_model=Project)
async def create_project(proj: ProjectCreate, user: dict = Depends(require_role(["admin"]))):
    now = get_ist_now_iso()
    
    # Check if project with same project_code already exists - merge if yes
    existing_project = await db.projects.find_one({"project_code": proj.project_code})
    
    if existing_project:
        # MERGE data with existing project
        
        # 1. Merge completed_hours (old + new)
        merged_hours = existing_project.get("completed_hours", 0) + proj.completed_hours
        
        # 2. Merge assigned_employees (combine unique employees)
        existing_employees = existing_project.get("assigned_employees", [])
        new_employees = proj.assigned_employees or []
        merged_employees = list(set(existing_employees + new_employees))
        
        # 3. Update end_date to new value (completion date)
        # 4. Keep earliest start_date
        existing_start = existing_project.get("start_date", proj.start_date)
        merged_start = min(existing_start, proj.start_date) if existing_start and proj.start_date else proj.start_date or existing_start
        
        update_data = {
            "name": proj.name or existing_project.get("name"),
            "type": proj.type or existing_project.get("type"),
            "start_date": merged_start,
            "end_date": proj.end_date,  # Use new end date
            "completed_hours": merged_hours,
            "assigned_employees": merged_employees,
            "status": proj.status or existing_project.get("status"),
            "client_username": proj.client_username or existing_project.get("client_username"),
            "scope_of_work": proj.scope_of_work or existing_project.get("scope_of_work"),
            "timesheet_link": proj.timesheet_link or existing_project.get("timesheet_link"),
            "updated_at": now
        }
        
        await db.projects.update_one(
            {"project_code": proj.project_code},
            {"$set": update_data}
        )
        updated = await db.projects.find_one({"project_code": proj.project_code}, {"_id": 0})
        return updated
    
    # Create new project if project_code doesn't exist
    proj_id = str(uuid.uuid4())
    proj_doc = {
        "id": proj_id,
        "name": proj.name,
        "type": proj.type,
        "project_code": proj.project_code,
        "start_date": proj.start_date,
        "end_date": proj.end_date or "",
        "completed_hours": proj.completed_hours,
        "assigned_employees": proj.assigned_employees,
        "status": proj.status,
        "client_username": proj.client_username,
        "scope_of_work": proj.scope_of_work,
        "timesheet_link": proj.timesheet_link,
        "is_late": proj.is_late or False,  # Manual late marking
        "created_at": now,
        "updated_at": now
    }
    
    await db.projects.insert_one(proj_doc)
    created_proj = await db.projects.find_one({"id": proj_id}, {"_id": 0})
    return created_proj

@api_router.put("/projects/{proj_id}", response_model=Project)
async def update_project(proj_id: str, proj: ProjectUpdate, user: dict = Depends(require_role(["admin"]))):
    now = get_ist_now_iso()
    update_data = {"updated_at": now}
    
    if proj.name:
        update_data["name"] = proj.name
    if proj.type:
        update_data["type"] = proj.type
    if proj.project_code:
        update_data["project_code"] = proj.project_code
    if proj.start_date:
        update_data["start_date"] = proj.start_date
    if proj.end_date is not None:
        update_data["end_date"] = proj.end_date
    if proj.assigned_employees is not None:
        update_data["assigned_employees"] = proj.assigned_employees
    if proj.status:
        update_data["status"] = proj.status
    if proj.client_username is not None:
        update_data["client_username"] = proj.client_username
    if proj.scope_of_work is not None:
        update_data["scope_of_work"] = proj.scope_of_work
    if proj.timesheet_link is not None:
        update_data["timesheet_link"] = proj.timesheet_link
    if proj.is_late is not None:
        update_data["is_late"] = proj.is_late  # Manual late marking
    
    result = await db.projects.update_one({"id": proj_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    
    updated = await db.projects.find_one({"id": proj_id}, {"_id": 0})
    return updated

@api_router.put("/projects/{proj_id}/late")
async def mark_project_late(proj_id: str, is_late: bool, user: dict = Depends(require_role(["admin"]))):
    result = await db.projects.update_one(
        {"id": proj_id},
        {"$set": {"is_late": is_late, "updated_at": get_ist_now_iso()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"message": "Project status updated"}

@api_router.get("/projects/{proj_id}/history")
async def get_project_history(proj_id: str, user: dict = Depends(get_current_user)):
    """Get complete work history for a project"""
    project = await db.projects.find_one({"id": proj_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Get all work entries for this project (by project_code)
    work_entries = await db.work_entries.find(
        {"$or": [
            {"project_code": project["project_code"]},
            {"project_id": project["id"]}  # Legacy support
        ]},
        {"_id": 0}
    ).sort("date", -1).to_list(10000)
    
    # Get employee details
    employees = await db.employees.find({}, {"_id": 0}).to_list(1000)
    emp_map = {emp["employee_id"]: emp for emp in employees}
    
    # Calculate developer-wise breakdown
    developer_hours = {}
    for entry in work_entries:
        emp_id = entry["employee_id"]
        if emp_id not in developer_hours:
            emp = emp_map.get(emp_id, {})
            developer_hours[emp_id] = {
                "employee_id": emp_id,
                "employee_name": emp.get("name", "Unknown"),
                "status": emp.get("status", "unknown"),
                "total_hours": 0,
                "entries_count": 0
            }
        developer_hours[emp_id]["total_hours"] += entry["hours"]
        developer_hours[emp_id]["entries_count"] += 1
    
    # Sort developers by hours descending
    developers = sorted(developer_hours.values(), key=lambda x: x["total_hours"], reverse=True)
    
    return {
        "project": project,
        "total_hours": sum(e["hours"] for e in work_entries),
        "total_developers": len(developer_hours),
        "active_developers": len([d for d in developers if d["status"] == "active"]),
        "developers": developers,
        "work_entries": work_entries[:100],  # Limit to last 100 entries
        "total_entries": len(work_entries)
    }

@api_router.delete("/projects/{proj_id}")
async def delete_project(proj_id: str, user: dict = Depends(require_role(["admin"]))):
    result = await db.projects.delete_one({"id": proj_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"message": "Project deleted"}

@api_router.delete("/projects")
async def delete_all_projects(user: dict = Depends(require_role(["admin"]))):
    """Delete ALL projects - Admin only"""
    result = await db.projects.delete_many({})
    return {"message": f"Deleted {result.deleted_count} projects", "deleted_count": result.deleted_count}

@api_router.get("/employee/is-pm")
async def check_if_pm_employee(user: dict = Depends(get_current_user)):
    """Check if current employee belongs to PM department"""
    if user["role"] != "employee":
        return {"is_pm": False}
    
    # Get employee details
    employee = await db.employees.find_one({"email": user["email"]})
    if not employee:
        return {"is_pm": False}
    
    # Get PM department ID
    pm_dept = await db.departments.find_one({"name": {"$regex": "^PM$", "$options": "i"}})
    if not pm_dept:
        return {"is_pm": False}
    
    pm_id = pm_dept.get("id")
    emp_depts = employee.get("department_ids", [])
    
    is_pm = pm_id in emp_depts
    return {
        "is_pm": is_pm,
        "employee_id": employee.get("employee_id"),
        "employee_name": employee.get("name")
    }

@api_router.get("/employee/is-team-leader")
async def check_if_team_leader(user: dict = Depends(get_current_user)):
    """Check if current employee is assigned as team leader to any employees"""
    if user["role"] != "employee":
        return {"is_team_leader": False, "team_members": []}
    
    # Get employee details
    employee = await db.employees.find_one({"email": user["email"]})
    if not employee:
        return {"is_team_leader": False, "team_members": []}
    
    emp_id = employee.get("employee_id")
    
    # Find all employees who have this employee as team leader
    team_members = await db.employees.find(
        {"team_leader_ids": emp_id},
        {"_id": 0, "employee_id": 1, "name": 1, "email": 1}
    ).to_list(1000)
    
    return {
        "is_team_leader": True,  # If we reach here, employee is in at least one project's team_leader_ids
        "employee_id": emp_id,
        "employee_name": employee.get("name"),
        "team_members": team_members
    }

@api_router.post("/projects/import")
async def import_projects(file: UploadFile = File(...), user: dict = Depends(require_role(["admin"]))):
    contents = await file.read()
    decoded = contents.decode('utf-8')
    csv_reader = csv.DictReader(io.StringIO(decoded))
    
    imported = 0
    merged = 0
    errors = []
    now = get_ist_now_iso()
    
    for row_num, row in enumerate(csv_reader, start=2):
        try:
            # Parse assigned employees (comma-separated)
            new_employees = [emp.strip() for emp in row.get('assigned_employees', '').split(',') if emp.strip()]
            new_hours = float(row.get('completed_hours', 0))
            
            # Check if project with same project_code exists
            existing_project = await db.projects.find_one({"project_code": row['project_code']})
            
            if existing_project:
                # MERGE with existing project
                # 1. Merge completed_hours (old + new)
                merged_hours = existing_project.get("completed_hours", 0) + new_hours
                
                # 2. Merge assigned_employees (combine unique)
                existing_employees = existing_project.get("assigned_employees", [])
                merged_employees = list(set(existing_employees + new_employees))
                
                # 3. Keep earliest start_date
                existing_start = existing_project.get("start_date", row['start_date'])
                merged_start = min(existing_start, row['start_date']) if existing_start and row['start_date'] else row['start_date'] or existing_start
                
                update_data = {
                    "name": row['name'] or existing_project.get("name"),
                    "type": row.get('type') or existing_project.get("type", "Development"),
                    "start_date": merged_start,
                    "end_date": row['end_date'],  # Use new end date
                    "completed_hours": merged_hours,
                    "assigned_employees": merged_employees,
                    "status": row.get('status') or existing_project.get("status", "active"),
                    "client_username": row.get('client_username') or existing_project.get("client_username", ""),
                    "scope_of_work": row.get('scope_of_work') or existing_project.get("scope_of_work", ""),
                    "timesheet_link": row.get('timesheet_link') or existing_project.get("timesheet_link", ""),
                    "updated_at": now
                }
                
                await db.projects.update_one(
                    {"project_code": row['project_code']},
                    {"$set": update_data}
                )
                merged += 1
            else:
                # Create new project
                proj_id = str(uuid.uuid4())
                proj_doc = {
                    "id": proj_id,
                    "name": row['name'],
                    "type": row.get('type', 'Development'),
                    "project_code": row['project_code'],
                    "start_date": row['start_date'],
                    "end_date": row['end_date'],
                    "completed_hours": new_hours,
                    "assigned_employees": new_employees,
                    "status": row.get('status', 'active'),
                    "client_username": row.get('client_username', ''),
                    "scope_of_work": row.get('scope_of_work', ''),
                    "timesheet_link": row.get('timesheet_link', ''),
                    "is_late": False,
                    "created_at": now,
                    "updated_at": now
                }
                await db.projects.insert_one(proj_doc)
                imported += 1
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
    
    return {
        "message": f"Imported {imported} new projects, merged {merged} existing projects",
        "imported": imported,
        "merged": merged,
        "errors": errors if errors else None
    }


# Work Entry Routes
@api_router.get("/work-entries")
async def get_work_entries(
    employee_id: Optional[str] = None,
    project_code: Optional[str] = None,
    date: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    
    # If project_code is provided, get all entries for that project (for team leader view)
    if project_code:
        query["project_code"] = project_code
        # If employee_id is also provided, filter by specific employee within that project
        if employee_id:
            query["employee_id"] = employee_id
    elif user["role"] == "employee":
        employee = await db.employees.find_one({"email": user["email"]})
        if not employee:
            raise HTTPException(status_code=404, detail="Employee profile not found. Please contact admin.")
        query["employee_id"] = employee["employee_id"]
    elif employee_id:
        query["employee_id"] = employee_id
    
    if date:
        query["date"] = date
    elif from_date or to_date:
        date_query = {}
        if from_date:
            date_query["$gte"] = from_date
        if to_date:
            date_query["$lte"] = to_date
        if date_query:
            query["date"] = date_query
    
    # Sort by date descending and increase limit for employees to see all their entries
    entries = await db.work_entries.find(query, {"_id": 0}).sort("date", -1).to_list(5000)
    return entries

@api_router.post("/work-entries", response_model=None)
async def create_work_entry(entry: WorkEntryCreate, user: dict = Depends(get_current_user)):
    employee = await db.employees.find_one({"email": user["email"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Validate project exists by project_code
    project = await db.projects.find_one({"project_code": entry.project_code})
    if not project:
        raise HTTPException(status_code=404, detail=f"Project with code '{entry.project_code}' not found")
    
    # Check if the date is a weekend or holiday
    from datetime import datetime
    entry_date = datetime.strptime(entry.date, "%Y-%m-%d")
    is_weekend = entry_date.weekday() in [5, 6]  # Saturday=5, Sunday=6
    
    # Check if it's a holiday
    is_holiday = False
    holiday_doc = await db.holidays.find_one({"date": entry.date})
    if holiday_doc:
        is_holiday = True
    
    now = get_ist_now_iso()
    
    # If weekend or holiday, create approval request instead of direct entry
    if is_weekend or is_holiday:
        approval_id = str(uuid.uuid4())
        approval_doc = {
            "id": approval_id,
            "employee_id": employee["employee_id"],
            "employee_name": employee["name"],
            "project_code": entry.project_code,
            "project_name": project.get("name", entry.project_code),
            "original_date": entry.date,
            "original_hours": entry.hours,
            "original_work_details": entry.work_details,
            "approved_date": None,
            "approved_hours": None,
            "status": "pending",
            "rejection_reason": "",
            "is_weekend": is_weekend,
            "is_holiday": is_holiday,
            "holiday_name": holiday_doc.get("name") if holiday_doc else None,
            "approved_by": None,
            "approved_at": None,
            "created_at": now
        }
        await db.weekend_approvals.insert_one(approval_doc)
        
        day_type = "Holiday" if is_holiday else ("Saturday" if entry_date.weekday() == 5 else "Sunday")
        return {
            "message": f"Work entry for {day_type} submitted for approval",
            "status": "pending_approval",
            "approval_id": approval_id
        }
    
    # Normal weekday flow - check if entry already exists
    existing_entry = await db.work_entries.find_one({
        "employee_id": employee["employee_id"],
        "date": entry.date,
        "project_code": entry.project_code
    })
    
    if existing_entry:
        # Update existing entry - add hours and append work details
        new_hours = existing_entry["hours"] + entry.hours
        new_details = existing_entry.get("work_details", "")
        if new_details and entry.work_details:
            new_details = f"{new_details}\n---\n{entry.work_details}"
        elif entry.work_details:
            new_details = entry.work_details
        
        await db.work_entries.update_one(
            {"id": existing_entry["id"]},
            {"$set": {
                "hours": new_hours,
                "work_details": new_details,
                "updated_at": now
            }}
        )
        
        # Update project completed hours (only add the new hours, not total)
        new_project_hours = project["completed_hours"] + entry.hours
        await db.projects.update_one(
            {"project_code": entry.project_code},
            {"$set": {"completed_hours": new_project_hours}}
        )
        
        updated_entry = await db.work_entries.find_one({"id": existing_entry["id"]}, {"_id": 0})
        return updated_entry
    else:
        # Create new entry
        entry_id = str(uuid.uuid4())
        
        entry_doc = {
            "id": entry_id,
            "employee_id": employee["employee_id"],
            "date": entry.date,
            "project_code": entry.project_code,
            "hours": entry.hours,
            "work_details": entry.work_details,
            "created_at": now
        }
        
        await db.work_entries.insert_one(entry_doc)
        
        # Update project completed hours
        new_hours = project["completed_hours"] + entry.hours
        await db.projects.update_one(
            {"project_code": entry.project_code},
            {"$set": {"completed_hours": new_hours}}
        )
        
        # Return without _id (MongoDB adds _id to entry_doc after insert)
        created_entry = await db.work_entries.find_one({"id": entry_id}, {"_id": 0})
        return created_entry

class AdminWorkEntryCreate(BaseModel):
    employee_id: str
    project_code: str  # Changed from project_id to project_code
    hours: float
    work_details: str
    date: str

@api_router.post("/work-entries/admin", response_model=WorkEntry)
async def create_admin_work_entry(entry: AdminWorkEntryCreate, user: dict = Depends(require_role(["admin", "hr"]))):
    # Validate employee exists
    employee = await db.employees.find_one({"employee_id": entry.employee_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Validate project exists by project_code
    project = await db.projects.find_one({"project_code": entry.project_code})
    if not project:
        raise HTTPException(status_code=404, detail=f"Project with code '{entry.project_code}' not found")
    
    # Check total hours for the day
    existing_entries = await db.work_entries.find({
        "employee_id": entry.employee_id,
        "date": entry.date
    }).to_list(1000)
    
    total_hours = sum(e["hours"] for e in existing_entries) + entry.hours
    
    entry_id = str(uuid.uuid4())
    now = get_ist_now_iso()
    
    entry_doc = {
        "id": entry_id,
        "employee_id": entry.employee_id,
        "date": entry.date,
        "project_code": entry.project_code,
        "hours": entry.hours,
        "work_details": entry.work_details,
        "created_at": now
    }
    
    await db.work_entries.insert_one(entry_doc)
    
    # Update project completed hours
    new_hours = project["completed_hours"] + entry.hours
    await db.projects.update_one(
        {"project_code": entry.project_code},
            {"$set": {"completed_hours": new_hours}}
        )
    
    # Return without _id
    created_entry = await db.work_entries.find_one({"id": entry_id}, {"_id": 0})
    return created_entry

@api_router.post("/work-entries/import")
async def import_work_entries(file: UploadFile = File(...), user: dict = Depends(require_role(["admin", "hr"]))):
    contents = await file.read()
    decoded = contents.decode('utf-8')
    csv_reader = csv.DictReader(io.StringIO(decoded))
    
    imported = 0
    errors = []
    now = get_ist_now_iso()
    
    for row_num, row in enumerate(csv_reader, start=2):
        try:
            # Validate employee
            employee = await db.employees.find_one({"employee_id": row["employee_id"]})
            if not employee:
                errors.append(f"Row {row_num}: Employee {row['employee_id']} not found")
                continue
            
            # Validate project by project_code
            project = await db.projects.find_one({"project_code": row["project_code"]})
            if not project:
                errors.append(f"Row {row_num}: Project with code '{row['project_code']}' not found")
                continue
            
            # Check if entry exists for same employee + date + project_code
            existing_entry = await db.work_entries.find_one({
                "employee_id": row["employee_id"],
                "date": row["date"],
                "project_code": row["project_code"]
            })
            
            if existing_entry:
                # Update existing entry - add hours
                new_hours = existing_entry["hours"] + float(row["hours"])
                new_details = existing_entry.get("work_details", "")
                if new_details and row.get("work_details"):
                    new_details = f"{new_details}\n---\n{row['work_details']}"
                elif row.get("work_details"):
                    new_details = row["work_details"]
                
                await db.work_entries.update_one(
                    {"id": existing_entry["id"]},
                    {"$set": {"hours": new_hours, "work_details": new_details, "updated_at": now}}
                )
            else:
                # Create new entry
                entry_id = str(uuid.uuid4())
                entry_doc = {
                    "id": entry_id,
                    "employee_id": row["employee_id"],
                    "date": row["date"],
                    "project_code": row["project_code"],
                    "hours": float(row["hours"]),
                    "work_details": row.get("work_details", ""),
                    "created_at": now
                }
                await db.work_entries.insert_one(entry_doc)
            
            # Update project hours
            new_project_hours = project["completed_hours"] + float(row["hours"])
            await db.projects.update_one(
                {"project_code": row["project_code"]},
                {"$set": {"completed_hours": new_project_hours}}
            )
            
            imported += 1
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
    
    return {
        "message": f"Imported {imported} work entries",
        "imported": imported,
        "errors": errors if errors else None
    }

class WorkEntryUpdate(BaseModel):
    project_code: Optional[str] = None  # Changed from project_id to project_code
    hours: Optional[float] = None
    work_details: Optional[str] = None
    date: Optional[str] = None

@api_router.put("/work-entries/{entry_id}")
async def update_work_entry(entry_id: str, update: WorkEntryUpdate, user: dict = Depends(require_role(["admin", "hr"]))):
    entry = await db.work_entries.find_one({"id": entry_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Work entry not found")
    
    old_hours = entry["hours"]
    old_project_code = entry.get("project_code") or entry.get("project_id", "")  # Support legacy data
    
    update_data = {"updated_at": get_ist_now_iso()}
    
    if update.project_code is not None:
        # Validate new project exists
        new_project = await db.projects.find_one({"project_code": update.project_code})
        if not new_project:
            raise HTTPException(status_code=404, detail=f"Project with code '{update.project_code}' not found")
        update_data["project_code"] = update.project_code
    if update.hours is not None:
        update_data["hours"] = update.hours
    if update.work_details is not None:
        update_data["work_details"] = update.work_details
    if update.date is not None:
        update_data["date"] = update.date
    
    await db.work_entries.update_one({"id": entry_id}, {"$set": update_data})
    
    # Update project hours if hours or project changed
    new_hours = update.hours if update.hours is not None else old_hours
    new_project_code = update.project_code if update.project_code is not None else old_project_code
    
    # Subtract from old project
    old_project = await db.projects.find_one({"project_code": old_project_code})
    if old_project:
        await db.projects.update_one(
            {"project_code": old_project_code},
            {"$inc": {"completed_hours": -old_hours}}
        )
    
    # Add to new project
    new_project = await db.projects.find_one({"project_code": new_project_code})
    if new_project:
        await db.projects.update_one(
            {"project_code": new_project_code},
            {"$inc": {"completed_hours": new_hours}}
        )
    
    updated = await db.work_entries.find_one({"id": entry_id}, {"_id": 0})
    return updated

@api_router.delete("/work-entries/{entry_id}")
async def delete_work_entry(entry_id: str, user: dict = Depends(require_role(["admin", "hr"]))):
    entry = await db.work_entries.find_one({"id": entry_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Work entry not found")
    
    # Update project hours - support both project_code and legacy project_id
    project_code = entry.get("project_code") or entry.get("project_id", "")
    project = await db.projects.find_one({"project_code": project_code})
    if project:
        await db.projects.update_one(
            {"project_code": project_code},
            {"$inc": {"completed_hours": -entry["hours"]}}
        )
    
    await db.work_entries.delete_one({"id": entry_id})
    return {"message": "Work entry deleted"}

@api_router.delete("/work-entries")
async def delete_all_work_entries(user: dict = Depends(require_role(["admin"]))):
    """Delete ALL work entries - Admin only"""
    result = await db.work_entries.delete_many({})
    # Also reset completed_hours on all projects
    await db.projects.update_many({}, {"$set": {"completed_hours": 0}})
    return {"message": f"Deleted {result.deleted_count} work entries", "deleted_count": result.deleted_count}

@api_router.delete("/work-entries/employee/{entry_id}")
async def delete_employee_work_entry(entry_id: str, user: dict = Depends(get_current_user)):
    """Employee can only delete their own work entries for current date"""
    entry = await db.work_entries.find_one({"id": entry_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Work entry not found")
    
    # Check if employee owns this entry
    if entry["employee_id"] != user.get("employee_id"):
        raise HTTPException(status_code=403, detail="You can only delete your own entries")
    
    # Check if entry is from today or future (IST)
    today_ist = get_ist_now().strftime("%Y-%m-%d")
    if entry["date"] < today_ist:
        raise HTTPException(status_code=403, detail="You can only delete entries from today or future dates")
    
    # Update project hours - support both project_code and legacy project_id
    project_code = entry.get("project_code") or entry.get("project_id", "")
    project = await db.projects.find_one({"project_code": project_code})
    if project:
        await db.projects.update_one(
            {"project_code": project_code},
            {"$inc": {"completed_hours": -entry["hours"]}}
        )
    
    await db.work_entries.delete_one({"id": entry_id})
    return {"message": "Work entry deleted"}

@api_router.put("/work-entries/employee/{entry_id}")
async def update_employee_work_entry(entry_id: str, update: WorkEntryUpdate, user: dict = Depends(get_current_user)):
    """Employee can only edit their own work entries for today or future dates"""
    entry = await db.work_entries.find_one({"id": entry_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Work entry not found")
    
    # Get employee record to verify ownership
    employee = await db.employees.find_one({"email": user["email"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee profile not found")
    
    # Check if employee owns this entry
    if entry["employee_id"] != employee["employee_id"]:
        raise HTTPException(status_code=403, detail="You can only edit your own entries")
    
    # Check if entry is from today or future (IST)
    today_ist = get_ist_now().strftime("%Y-%m-%d")
    if entry["date"] < today_ist:
        raise HTTPException(status_code=403, detail="You can only edit entries from today or future dates")
    
    old_hours = entry["hours"]
    old_project_code = entry.get("project_code") or entry.get("project_id", "")
    
    update_data = {"updated_at": get_ist_now_iso()}
    
    if update.project_code is not None:
        # Validate new project exists and employee is assigned
        new_project = await db.projects.find_one({"project_code": update.project_code})
        if not new_project:
            raise HTTPException(status_code=404, detail=f"Project with code '{update.project_code}' not found")
        update_data["project_code"] = update.project_code
    if update.hours is not None:
        update_data["hours"] = update.hours
    if update.work_details is not None:
        update_data["work_details"] = update.work_details
    
    await db.work_entries.update_one({"id": entry_id}, {"$set": update_data})
    
    # Update project hours if hours or project changed
    new_hours = update.hours if update.hours is not None else old_hours
    new_project_code = update.project_code if update.project_code is not None else old_project_code
    
    # Subtract from old project
    if old_project_code:
        await db.projects.update_one(
            {"project_code": old_project_code},
            {"$inc": {"completed_hours": -old_hours}}
        )
    
    # Add to new project
    if new_project_code:
        await db.projects.update_one(
            {"project_code": new_project_code},
            {"$inc": {"completed_hours": new_hours}}
        )
    
    updated = await db.work_entries.find_one({"id": entry_id}, {"_id": 0})
    return updated


@api_router.get("/work-entries/summary")
async def get_work_summary(user: dict = Depends(require_role(["admin", "hr"]))):
    # Get all projects in one query
    projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    
    # Get ALL work entries in ONE query and group by project_code
    all_entries = await db.work_entries.find({}, {"_id": 0, "project_code": 1, "employee_id": 1, "hours": 1}).to_list(100000)
    
    # Build a map of project_code -> employee_hours
    project_hours_map = {}
    for entry in all_entries:
        pc = entry.get("project_code")
        if not pc:
            continue
        if pc not in project_hours_map:
            project_hours_map[pc] = {}
        emp_id = entry["employee_id"]
        if emp_id not in project_hours_map[pc]:
            project_hours_map[pc][emp_id] = 0
        project_hours_map[pc][emp_id] += entry["hours"]
    
    # Build summary using the pre-computed map
    summary = []
    for project in projects:
        project_code = project["project_code"]
        employee_hours = project_hours_map.get(project_code, {})
        
        summary.append({
            "project_id": project["id"],
            "project_name": project["name"],
            "project_code": project_code,
            "status": project["status"],
            "completed_hours": project["completed_hours"],
            "assigned_employees": project["assigned_employees"],
            "employee_hours": employee_hours
        })
    
    return summary

@api_router.get("/work-entries/detailed-summary")
async def get_detailed_work_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    department_id: Optional[str] = None,
    employee_id: Optional[str] = None,
    project_code: Optional[str] = None,
    employee_status: Optional[str] = "active",  # Default to active employees
    page: int = 1,
    page_size: int = 50,
    user: dict = Depends(require_role(["admin", "hr"]))
):
    # Build query
    query = {}
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["date"] = {"$gte": start_date}
    elif end_date:
        query["date"] = {"$lte": end_date}
    
    if employee_id:
        query["employee_id"] = employee_id
    
    # Filter by project_code at database level
    if project_code:
        query["project_code"] = project_code
    
    # Get all work entries
    entries = await db.work_entries.find(query, {"_id": 0}).to_list(50000)
    
    # Get all employees for filtering
    employees = await db.employees.find({}, {"_id": 0}).to_list(1000)
    employee_map = {emp["employee_id"]: emp for emp in employees}
    
    # Filter by employee status (active/inactive/all)
    if employee_status and employee_status != "all":
        entries = [e for e in entries if employee_map.get(e["employee_id"], {}).get("status") == employee_status]
    
    # Filter by department if specified
    if department_id:
        # Get the department to find its name (in case there are duplicates)
        dept = await db.departments.find_one({"id": department_id})
        dept_name = dept.get("name") if dept else None
        
        # Find all department IDs with the same name (to handle duplicates)
        matching_dept_ids = set()
        if dept_name:
            async for d in db.departments.find({"name": dept_name}, {"id": 1}):
                matching_dept_ids.add(d["id"])
        else:
            matching_dept_ids.add(department_id)
        
        # Check both department_id (legacy) and department_ids (multiple departments)
        entries = [e for e in entries if 
            employee_map.get(e["employee_id"], {}).get("department_id") in matching_dept_ids or
            any(did in matching_dept_ids for did in employee_map.get(e["employee_id"], {}).get("department_ids", []))]
    
    # Get all projects
    projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    project_map = {proj["project_code"]: proj for proj in projects}
    
    # Group by employee and date
    grouped = {}
    for entry in entries:
        key = (entry["employee_id"], entry["date"])
        if key not in grouped:
            emp = employee_map.get(entry["employee_id"], {})
            grouped[key] = {
                "employee_id": entry["employee_id"],
                "employee_name": emp.get("name", "Unknown"),
                "employee_status": emp.get("status", "unknown"),
                "department_id": emp.get("department_id", ""),
                "date": entry["date"],
                "projects": [],
                "total_hours": 0
            }
        
        # Support both project_code (new) and project_id (legacy)
        proj_code = entry.get("project_code") or entry.get("project_id", "")
        project = project_map.get(proj_code, {})
        grouped[key]["projects"].append({
            "id": entry.get("id", ""),  # Work entry ID for deletion
            "project_id": project.get("id", ""),
            "project_code": proj_code,
            "project_name": project.get("name", proj_code or "Unknown"),
            "hours": entry["hours"],
            "work_details": entry.get("work_details", ""),
            "is_compensation": entry.get("is_compensation", False)
        })
        grouped[key]["total_hours"] += entry["hours"]
    
    # Convert to list and sort
    summary = list(grouped.values())
    summary.sort(key=lambda x: (x["date"], x["employee_name"]), reverse=True)
    
    # Pagination
    total_count = len(summary)
    total_pages = (total_count + page_size - 1) // page_size
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    paginated_summary = summary[start_idx:end_idx]
    
    return {
        "data": paginated_summary,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total_count": total_count,
            "total_pages": total_pages
        }
    }

@api_router.get("/work-entries/export")
async def export_work_entries(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(require_role(["admin", "hr"]))
):
    from fastapi.responses import StreamingResponse
    
    query = {}
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    
    entries = await db.work_entries.find(query, {"_id": 0}).to_list(10000)
    
    # Get employee and project details
    employees = await db.employees.find({}, {"_id": 0}).to_list(1000)
    projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    
    employee_map = {emp["employee_id"]: emp["name"] for emp in employees}
    project_map = {proj["project_code"]: proj["name"] for proj in projects}
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Employee ID", "Employee Name", "Project Code", "Project Name", "Hours", "Work Details"])
    
    for entry in entries:
        proj_code = entry.get("project_code") or entry.get("project_id", "")
        writer.writerow([
            entry["date"],
            entry["employee_id"],
            employee_map.get(entry["employee_id"], "Unknown"),
            proj_code,
            project_map.get(proj_code, "Unknown"),
            entry["hours"],
            entry.get("work_details", "")
        ])
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=work_entries_export.csv"}
    )


# Weekend/Holiday Approval Routes
@api_router.get("/weekend-approvals")
async def get_weekend_approvals(
    status: Optional[str] = "pending",
    month: Optional[str] = None,  # Format: YYYY-MM
    user: dict = Depends(require_role(["admin", "hr"]))
):
    """Get weekend/holiday approval requests"""
    query = {}
    if status and status != "all":
        query["status"] = status
    
    if month:
        # Filter by month (YYYY-MM)
        query["original_date"] = {"$regex": f"^{month}"}
    
    approvals = []
    async for approval in db.weekend_approvals.find(query, {"_id": 0}).sort("created_at", -1):
        approvals.append(approval)
    
    return approvals

@api_router.get("/weekend-approvals/pending-count")
async def get_pending_approval_count(user: dict = Depends(require_role(["admin", "hr"]))):
    """Get count of pending weekend/holiday approvals"""
    count = await db.weekend_approvals.count_documents({"status": "pending"})
    return {"count": count}

@api_router.get("/weekend-approvals/history")
async def get_approval_history(
    month: Optional[str] = None,  # Format: YYYY-MM
    employee_id: Optional[str] = None,
    user: dict = Depends(require_role(["admin", "hr"]))
):
    """Get approval history with filters - starting from March 2026"""
    query = {"status": {"$in": ["approved", "rejected"]}}
    
    if month:
        query["original_date"] = {"$regex": f"^{month}"}
    else:
        # Default: show from March 2026 onwards
        query["original_date"] = {"$gte": "2026-03-01"}
    
    if employee_id:
        query["employee_id"] = employee_id
    
    history = []
    async for record in db.weekend_approvals.find(query, {"_id": 0}).sort("approved_at", -1):
        history.append(record)
    
    return history

@api_router.get("/weekend-approvals/my-pending")
async def get_my_pending_approvals(user: dict = Depends(get_current_user)):
    """Get employee's pending weekend/holiday approval requests"""
    employee = await db.employees.find_one({"email": user["email"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    requests = []
    async for req in db.weekend_approvals.find(
        {"employee_id": employee["employee_id"], "status": "pending"},
        {"_id": 0}
    ).sort("created_at", -1):
        requests.append(req)
    
    return requests

@api_router.get("/weekend-approvals/{approval_id}")
async def get_approval_detail(approval_id: str, user: dict = Depends(require_role(["admin", "hr"]))):
    """Get single approval detail"""
    approval = await db.weekend_approvals.find_one({"id": approval_id}, {"_id": 0})
    if not approval:
        raise HTTPException(status_code=404, detail="Approval request not found")
    return approval

@api_router.put("/weekend-approvals/{approval_id}/approve")
async def approve_weekend_entry(
    approval_id: str,
    action: WeekendApprovalAction,
    user: dict = Depends(require_role(["admin", "hr"]))
):
    """Approve weekend/holiday work entry with optional edits"""
    approval = await db.weekend_approvals.find_one({"id": approval_id})
    if not approval:
        raise HTTPException(status_code=404, detail="Approval request not found")
    
    if approval["status"] != "pending":
        raise HTTPException(status_code=400, detail="This request has already been processed")
    
    now = get_ist_now_iso()
    
    # Use edited values or original values
    final_date = action.approved_date if action.approved_date else approval["original_date"]
    final_hours = action.approved_hours if action.approved_hours is not None else approval["original_hours"]
    
    # Update approval record
    await db.weekend_approvals.update_one(
        {"id": approval_id},
        {"$set": {
            "status": "approved",
            "approved_date": final_date,
            "approved_hours": final_hours,
            "approved_by": user["username"],
            "approved_at": now,
            "is_compensation": action.is_compensation or False
        }}
    )
    
    # Create actual work entry with approved values
    entry_id = str(uuid.uuid4())
    entry_doc = {
        "id": entry_id,
        "employee_id": approval["employee_id"],
        "date": final_date,
        "project_code": approval["project_code"],
        "hours": final_hours,
        "work_details": approval["original_work_details"],
        "created_at": now,
        "from_weekend_approval": True,
        "approval_id": approval_id,
        "is_compensation": action.is_compensation or False
    }
    
    await db.work_entries.insert_one(entry_doc)
    
    # Update project completed hours
    project = await db.projects.find_one({"project_code": approval["project_code"]})
    if project:
        new_hours = project.get("completed_hours", 0) + final_hours
        await db.projects.update_one(
            {"project_code": approval["project_code"]},
            {"$set": {"completed_hours": new_hours}}
        )
    
    return {
        "message": "Work entry approved and added to records",
        "work_entry_id": entry_id,
        "approved_date": final_date,
        "approved_hours": final_hours
    }

@api_router.put("/weekend-approvals/{approval_id}/reject")
async def reject_weekend_entry(
    approval_id: str,
    action: WeekendApprovalAction,
    user: dict = Depends(require_role(["admin", "hr"]))
):
    """Reject weekend/holiday work entry"""
    approval = await db.weekend_approvals.find_one({"id": approval_id})
    if not approval:
        raise HTTPException(status_code=404, detail="Approval request not found")
    
    if approval["status"] != "pending":
        raise HTTPException(status_code=400, detail="This request has already been processed")
    
    if not action.rejection_reason:
        raise HTTPException(status_code=400, detail="Rejection reason is required")
    
    now = get_ist_now_iso()
    
    await db.weekend_approvals.update_one(
        {"id": approval_id},
        {"$set": {
            "status": "rejected",
            "rejection_reason": action.rejection_reason,
            "approved_by": user["username"],
            "approved_at": now
        }}
    )
    
    return {"message": "Work entry rejected", "reason": action.rejection_reason}

@api_router.get("/weekend-approvals/employee/my-requests")
async def get_my_approval_requests(user: dict = Depends(get_current_user)):
    """Get employee's own weekend/holiday approval requests"""
    employee = await db.employees.find_one({"email": user["email"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    requests = []
    async for req in db.weekend_approvals.find(
        {"employee_id": employee["employee_id"]},
        {"_id": 0}
    ).sort("created_at", -1):
        requests.append(req)
    
    return requests

@api_router.put("/weekend-approvals/employee/{approval_id}")
async def update_my_pending_approval(approval_id: str, update: WorkEntryUpdate, user: dict = Depends(get_current_user)):
    """Employee can edit their own pending approval requests"""
    employee = await db.employees.find_one({"email": user["email"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    approval = await db.weekend_approvals.find_one({"id": approval_id})
    if not approval:
        raise HTTPException(status_code=404, detail="Approval request not found")
    
    if approval["employee_id"] != employee["employee_id"]:
        raise HTTPException(status_code=403, detail="You can only edit your own requests")
    
    if approval["status"] != "pending":
        raise HTTPException(status_code=400, detail="Can only edit pending requests")
    
    update_data = {"updated_at": get_ist_now_iso()}
    if update.hours is not None:
        update_data["original_hours"] = update.hours
    if update.work_details is not None:
        update_data["work_details"] = update.work_details
    
    await db.weekend_approvals.update_one({"id": approval_id}, {"$set": update_data})
    
    updated = await db.weekend_approvals.find_one({"id": approval_id}, {"_id": 0})
    return updated

@api_router.delete("/weekend-approvals/employee/{approval_id}")
async def delete_my_approval(approval_id: str, user: dict = Depends(get_current_user)):
    """Employee can delete their own approval requests (pending, approved, or rejected)"""
    employee = await db.employees.find_one({"email": user["email"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    approval = await db.weekend_approvals.find_one({"id": approval_id})
    if not approval:
        raise HTTPException(status_code=404, detail="Approval request not found")
    
    if approval["employee_id"] != employee["employee_id"]:
        raise HTTPException(status_code=403, detail="You can only delete your own requests")
    
    now = get_ist_now_iso()
    
    # For approved entries: also delete the corresponding work_entry and update project hours
    if approval["status"] == "approved":
        work_entry = await db.work_entries.find_one({"approval_id": approval_id})
        if work_entry:
            # Deduct hours from project before deleting
            project = await db.projects.find_one({"project_code": work_entry["project_code"]})
            if project:
                new_hours = max(0, project.get("completed_hours", 0) - work_entry["hours"])
                await db.projects.update_one(
                    {"project_code": work_entry["project_code"]},
                    {"$set": {"completed_hours": new_hours}}
                )
            # Delete the work entry
            await db.work_entries.delete_one({"approval_id": approval_id})
    
    # For pending entries: just delete
    if approval["status"] == "pending":
        await db.weekend_approvals.delete_one({"id": approval_id})
        return {"message": "Pending request deleted"}
    
    # For approved/rejected: soft delete for audit trail
    await db.weekend_approvals.update_one(
        {"id": approval_id},
        {"$set": {
            "status": "deleted",
            "deleted_by": employee["employee_id"],
            "deleted_at": now
        }}
    )
    
    return {"message": "Entry deleted from history"}

@api_router.delete("/weekend-approvals/{approval_id}")
async def delete_approved_entry(
    approval_id: str,
    user: dict = Depends(require_role(["admin"]))  # Admin only
):
    """Delete an approved or rejected weekend/holiday entry from history"""
    # Find the approval record
    approval = await db.weekend_approvals.find_one({"id": approval_id})
    if not approval:
        raise HTTPException(status_code=404, detail="Approval record not found")
    
    if approval["status"] not in ["approved", "rejected"]:
        raise HTTPException(status_code=400, detail="Only approved or rejected entries can be deleted from history")
    
    now = get_ist_now_iso()
    
    # For approved entries: delete the corresponding work_entry
    if approval["status"] == "approved":
        work_entry = await db.work_entries.find_one({"approval_id": approval_id})
        if work_entry:
            # Deduct hours from project before deleting
            project = await db.projects.find_one({"project_code": work_entry["project_code"]})
            if project:
                new_hours = max(0, project.get("completed_hours", 0) - work_entry["hours"])
                await db.projects.update_one(
                    {"project_code": work_entry["project_code"]},
                    {"$set": {"completed_hours": new_hours}}
                )
            
            # Delete the work entry
            await db.work_entries.delete_one({"approval_id": approval_id})
    
    # Update approval record with deletion info (soft delete for audit)
    await db.weekend_approvals.update_one(
        {"id": approval_id},
        {"$set": {
            "status": "deleted",
            "deleted_by": user["username"],
            "deleted_at": now
        }}
    )
    
    return {
        "message": "Entry deleted successfully from both history and work entries",
        "deleted_approval_id": approval_id,
        "work_entry_deleted": work_entry is not None
    }


# Email Configuration Routes
@api_router.get("/email-config")
async def get_email_config(user: dict = Depends(require_role(["admin"]))):
    """Get email configuration (admin only)"""
    config = await db.email_config.find_one({}, {"_id": 0})
    if not config:
        # Return default config
        return {
            "smtp_host": "smtp.gmail.com",
            "smtp_port": 587,
            "smtp_email": "hr.zestbrains@gmail.com",
            "smtp_password": "",
            "enable_ssl": True,
            "cc_emails": "",
            "is_enabled": False
        }
    # Hide password in response
    config["smtp_password"] = "••••••••" if config.get("smtp_password") else ""
    return config

@api_router.put("/email-config")
async def update_email_config(config: EmailConfigUpdate, user: dict = Depends(require_role(["admin"]))):
    """Update email configuration (admin only)"""
    now = get_ist_now_iso()
    
    # Get existing config to preserve password if not provided
    existing = await db.email_config.find_one({})
    
    update_data = {
        "smtp_host": config.smtp_host,
        "smtp_port": config.smtp_port,
        "smtp_email": config.smtp_email,
        "enable_ssl": config.enable_ssl,
        "cc_emails": config.cc_emails,
        "is_enabled": config.is_enabled,
        "updated_at": now,
        "updated_by": user["username"]
    }
    
    # Only update password if provided (not empty and not placeholder)
    if config.smtp_password and config.smtp_password != "••••••••":
        update_data["smtp_password"] = config.smtp_password
    elif existing and existing.get("smtp_password"):
        update_data["smtp_password"] = existing["smtp_password"]
    
    await db.email_config.update_one(
        {},
        {"$set": update_data},
        upsert=True
    )
    
    return {"message": "Email configuration updated successfully"}

@api_router.post("/email-config/test")
async def test_email_config(user: dict = Depends(require_role(["admin"]))):
    """Send test email to verify SMTP configuration"""
    config = await db.email_config.find_one({}, {"_id": 0})
    if not config or not config.get("smtp_password"):
        raise HTTPException(status_code=400, detail="Email configuration not complete")
    
    try:
        smtp_host = config.get("smtp_host", "smtp.gmail.com")
        smtp_port = config.get("smtp_port", 587)
        smtp_email = config.get("smtp_email")
        smtp_password = config.get("smtp_password")
        enable_ssl = config.get("enable_ssl", True)
        
        # Create test message
        msg = MIMEMultipart('alternative')
        msg['Subject'] = "Test Email - Zestbrains HR Portal"
        msg['From'] = smtp_email
        msg['To'] = smtp_email
        
        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif;">
            <h2>Test Email</h2>
            <p>This is a test email from Zestbrains HR Portal.</p>
            <p>If you received this email, your SMTP configuration is working correctly.</p>
            <p>Sent at: {get_ist_now().strftime("%d %b %Y, %I:%M %p IST")}</p>
        </body>
        </html>
        """
        msg.attach(MIMEText(html_body, 'html'))
        
        if enable_ssl:
            context = ssl.create_default_context()
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls(context=context)
                server.login(smtp_email, smtp_password)
                server.sendmail(smtp_email, smtp_email, msg.as_string())
        else:
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.login(smtp_email, smtp_password)
                server.sendmail(smtp_email, smtp_email, msg.as_string())
        
        return {"message": "Test email sent successfully! Check your inbox."}
    
    except smtplib.SMTPAuthenticationError:
        raise HTTPException(status_code=400, detail="SMTP Authentication failed. Check your email and password/app password.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send test email: {str(e)}")


# Leave Routes
@api_router.post("/leaves/apply")
async def apply_leave(leave: LeaveApplication, user: dict = Depends(get_current_user)):
    employee = await db.employees.find_one({"email": user["email"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    leave_id = str(uuid.uuid4())
    now = get_ist_now_iso()
    
    # Calculate days based on leave_dates if provided
    if leave.leave_dates and len(leave.leave_dates) > 0:
        # Sum up days: full = 1, first_half/second_half = 0.5
        days_count = sum(
            0.5 if ld.day_type in ['half', 'first_half', 'second_half'] else 1 
            for ld in leave.leave_dates
        )
        # Store the leave dates info for display
        leave_dates_info = [{"date": ld.date, "day_type": ld.day_type} for ld in leave.leave_dates]
    else:
        # Fallback to date range calculation
        from_date = datetime.fromisoformat(leave.from_date)
        to_date = datetime.fromisoformat(leave.to_date)
        days_count = (to_date - from_date).days + 1
        leave_dates_info = None
    
    leave_doc = {
        "id": leave_id,
        "employee_id": employee["employee_id"],
        "from_date": leave.from_date,
        "to_date": leave.to_date,
        "leave_type": leave.leave_type,
        "reason": leave.reason,
        "status": "pending",
        "days_count": days_count,
        "leave_dates_input": leave_dates_info,  # Store employee's half/full day selections
        "approved_by": None,
        "approved_date": None,
        "comments": "",
        "created_at": now
    }
    
    await db.leave_applications.insert_one(leave_doc)
    # Return without _id
    created_leave = await db.leave_applications.find_one({"id": leave_id}, {"_id": 0})
    return created_leave

class LeaveApplicationUpdate(BaseModel):
    from_date: str
    to_date: str
    reason: str

@api_router.put("/leaves/applications/{leave_id}")
async def update_leave_application(leave_id: str, leave_update: LeaveApplicationUpdate, user: dict = Depends(get_current_user)):
    """Update a pending leave application (only before admin action)"""
    employee = await db.employees.find_one({"email": user["email"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    leave_app = await db.leave_applications.find_one({"id": leave_id})
    if not leave_app:
        raise HTTPException(status_code=404, detail="Leave application not found")
    
    # Verify ownership
    if leave_app["employee_id"] != employee["employee_id"]:
        raise HTTPException(status_code=403, detail="Not authorized to update this application")
    
    # Only allow update if status is pending
    if leave_app["status"] != "pending":
        raise HTTPException(status_code=400, detail="Cannot edit - application has already been processed")
    
    # Calculate new days
    from_date = datetime.fromisoformat(leave_update.from_date)
    to_date = datetime.fromisoformat(leave_update.to_date)
    days_count = (to_date - from_date).days + 1
    
    now = get_ist_now_iso()
    
    await db.leave_applications.update_one(
        {"id": leave_id},
        {"$set": {
            "from_date": leave_update.from_date,
            "to_date": leave_update.to_date,
            "reason": leave_update.reason,
            "days_count": days_count,
            "updated_at": now
        }}
    )
    
    updated = await db.leave_applications.find_one({"id": leave_id}, {"_id": 0})
    return updated

@api_router.delete("/leaves/applications/{leave_id}")
async def delete_leave_application(leave_id: str, user: dict = Depends(get_current_user)):
    """Delete a pending leave application (only before admin action)"""
    employee = await db.employees.find_one({"email": user["email"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    leave_app = await db.leave_applications.find_one({"id": leave_id})
    if not leave_app:
        raise HTTPException(status_code=404, detail="Leave application not found")
    
    # Verify ownership
    if leave_app["employee_id"] != employee["employee_id"]:
        raise HTTPException(status_code=403, detail="Not authorized to delete this application")
    
    # Only allow delete if status is pending
    if leave_app["status"] != "pending":
        raise HTTPException(status_code=400, detail="Cannot delete - application has already been processed")
    
    await db.leave_applications.delete_one({"id": leave_id})
    return {"message": "Leave application deleted successfully"}

# Admin/HR Leave Management Routes
class AdminLeaveApplicationUpdate(BaseModel):
    from_date: str
    to_date: str
    reason: str
    status: Optional[str] = None
    leave_dates: Optional[List[LeaveDateType]] = []
    comments: Optional[str] = ""
    reject_reason: Optional[str] = ""

@api_router.put("/leaves/applications/{leave_id}/admin-edit")
async def admin_edit_leave_application(leave_id: str, leave_update: AdminLeaveApplicationUpdate, user: dict = Depends(require_role(["admin", "hr"]))):
    """Admin/HR can edit any leave application (pending, approved, or rejected)"""
    leave_app = await db.leave_applications.find_one({"id": leave_id})
    if not leave_app:
        raise HTTPException(status_code=404, detail="Leave application not found")
    
    employee_id = leave_app["employee_id"]
    old_status = leave_app.get("status", "pending")
    old_leave_dates = leave_app.get("leave_dates", [])
    new_leave_dates = [ld.dict() for ld in leave_update.leave_dates] if leave_update.leave_dates else old_leave_dates
    new_status = leave_update.status if leave_update.status else old_status
    
    # Calculate leave balance changes if status is approved or was approved
    if old_status == "approved" or new_status == "approved":
        # Calculate old leave totals (only if was approved)
        old_pl = 0.0
        old_cl = 0.0
        if old_status == "approved":
            for ld in old_leave_dates:
                lt = ld.get("leave_type", "")
                if lt.lower() == "rejected":
                    continue
                if lt == "PL/2 & CL/2":
                    old_pl += 0.5
                    old_cl += 0.5
                elif "PL" in lt:
                    old_pl += 0.5 if "Half" in lt else 1.0
                else:
                    old_cl += 0.5 if "Half" in lt else 1.0
        
        # Calculate new leave totals (only if will be approved)
        new_pl = 0.0
        new_cl = 0.0
        if new_status == "approved":
            for ld in new_leave_dates:
                lt = ld.get("leave_type", "")
                if lt.lower() == "rejected":
                    continue
                if lt == "PL/2 & CL/2":
                    new_pl += 0.5
                    new_cl += 0.5
                elif "PL" in lt:
                    new_pl += 0.5 if "Half" in lt else 1.0
                else:
                    new_cl += 0.5 if "Half" in lt else 1.0
        
        # Calculate difference and update employee leave balance
        pl_diff = new_pl - old_pl
        cl_diff = new_cl - old_cl
        
        if pl_diff != 0 or cl_diff != 0:
            update_balance = {}
            if pl_diff != 0:
                update_balance["pl_taken"] = pl_diff
            if cl_diff != 0:
                update_balance["cl_taken"] = cl_diff
            
            await db.employees.update_one(
                {"employee_id": employee_id},
                {"$inc": update_balance}
            )
        
        # Handle leave_records: Delete old and create new if approved
        old_dates_set = {ld["date"] for ld in old_leave_dates if ld.get("leave_type", "").lower() != "rejected"}
        new_dates_set = {ld["date"] for ld in new_leave_dates if ld.get("leave_type", "").lower() != "rejected"}
        
        # Delete records for removed dates
        dates_to_remove = old_dates_set - new_dates_set
        if dates_to_remove and old_status == "approved":
            await db.leave_records.delete_many({
                "employee_id": employee_id,
                "date": {"$in": list(dates_to_remove)}
            })
        
        # Add/update records for new dates if approved
        if new_status == "approved":
            for ld in new_leave_dates:
                lt = ld.get("leave_type", "")
                if lt.lower() != "rejected":
                    # Calculate leave_days based on leave type
                    if "Half" in lt or "/2" in lt:
                        leave_days = 0.5
                    else:
                        leave_days = 1.0
                    
                    # Determine if PL or CL for the record type
                    record_type = "PL" if "PL" in lt else "CL"
                    
                    await db.leave_records.update_one(
                        {"employee_id": employee_id, "date": ld["date"]},
                        {"$set": {
                            "leave_type": record_type,
                            "leave_days": leave_days,
                            "application_id": leave_id,
                            "status": "Taken",
                            "updated_at": get_ist_now_iso()
                        }},
                        upsert=True
                    )
    
    # Calculate new days count based on leave_dates
    if leave_update.leave_dates:
        days_count = len([ld for ld in new_leave_dates if ld.get("leave_type", "").lower() != "rejected"])
        # Sum up actual day values (half days = 0.5)
        actual_days = 0
        for ld in new_leave_dates:
            lt = ld.get("leave_type", "")
            if lt.lower() == "rejected":
                continue
            if "Half" in lt or "PL/2" in lt or "CL/2" in lt:
                actual_days += 0.5
            else:
                actual_days += 1
    else:
        from_date = datetime.fromisoformat(leave_update.from_date)
        to_date = datetime.fromisoformat(leave_update.to_date)
        days_count = (to_date - from_date).days + 1
        actual_days = days_count
    
    now = get_ist_now_iso()
    
    update_data = {
        "from_date": leave_update.from_date,
        "to_date": leave_update.to_date,
        "reason": leave_update.reason,
        "days_count": actual_days,
        "updated_at": now,
        "last_modified_by": user["username"]
    }
    
    if leave_update.status:
        update_data["status"] = leave_update.status
    if leave_update.leave_dates:
        update_data["leave_dates"] = new_leave_dates
    if leave_update.comments:
        update_data["comments"] = leave_update.comments
    if leave_update.reject_reason:
        update_data["reject_reason"] = leave_update.reject_reason
    
    await db.leave_applications.update_one(
        {"id": leave_id},
        {"$set": update_data}
    )
    
    updated = await db.leave_applications.find_one({"id": leave_id}, {"_id": 0})
    return updated

@api_router.delete("/leaves/applications/{leave_id}/admin-delete")
async def admin_delete_leave_application(leave_id: str, user: dict = Depends(require_role(["admin", "hr"]))):
    """Admin/HR can delete any leave application (pending, approved, or rejected)"""
    leave_app = await db.leave_applications.find_one({"id": leave_id})
    if not leave_app:
        raise HTTPException(status_code=404, detail="Leave application not found")
    
    # If leave was approved, we need to reverse the leave balance
    if leave_app.get("status") == "approved" and leave_app.get("leave_dates"):
        total_pl = 0.0
        total_cl = 0.0
        
        for ld in leave_app["leave_dates"]:
            leave_type = ld.get("leave_type", "")
            if leave_type.lower() == "rejected":
                continue
            
            if leave_type == "PL/2 & CL/2":
                total_pl += 0.5
                total_cl += 0.5
            elif "PL" in leave_type:
                total_pl += 0.5 if "Half" in leave_type else 1.0
            else:
                total_cl += 0.5 if "Half" in leave_type else 1.0
        
        # Reverse the leave balance
        if total_pl > 0:
            await db.employees.update_one(
                {"employee_id": leave_app["employee_id"]},
                {"$inc": {"pl_taken": -total_pl}}
            )
        if total_cl > 0:
            await db.employees.update_one(
                {"employee_id": leave_app["employee_id"]},
                {"$inc": {"cl_taken": -total_cl}}
            )
        
        # Delete associated leave records
        await db.leave_records.delete_many({
            "employee_id": leave_app["employee_id"],
            "date": {"$in": [ld["date"] for ld in leave_app["leave_dates"] if ld.get("leave_type", "").lower() != "rejected"]}
        })
    
    await db.leave_applications.delete_one({"id": leave_id})
    return {"message": "Leave application deleted successfully"}

@api_router.get("/leaves/applications")
async def get_leave_applications(
    status: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    
    if user["role"] == "employee":
        employee = await db.employees.find_one({"email": user["email"]})
        if employee:
            query["employee_id"] = employee["employee_id"]
    
    if status:
        query["status"] = status
    
    applications = await db.leave_applications.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return applications

@api_router.get("/leaves/history")
async def get_leave_history(
    employee_id: Optional[str] = None,
    status: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    user: dict = Depends(require_role(["admin", "hr"]))
):
    """Get leave application history with filters for Admin/HR"""
    query = {}
    
    if employee_id:
        query["employee_id"] = employee_id
    
    if status:
        query["status"] = status
    
    # Filter by date range - applications whose dates overlap with the filter range
    if from_date or to_date:
        date_query = {}
        if from_date:
            date_query["$gte"] = from_date
        if to_date:
            date_query["$lte"] = to_date
        
        if from_date and to_date:
            # Applications where from_date is within range OR to_date is within range
            query["$or"] = [
                {"from_date": date_query},
                {"to_date": date_query},
                {"$and": [{"from_date": {"$lte": from_date}}, {"to_date": {"$gte": to_date}}]}
            ]
        elif from_date:
            query["to_date"] = {"$gte": from_date}
        elif to_date:
            query["from_date"] = {"$lte": to_date}
    
    applications = await db.leave_applications.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return applications

@api_router.get("/leaves/my-details")
async def get_my_leave_details(user: dict = Depends(get_current_user)):
    """Get leave details for the currently logged-in employee"""
    from collections import defaultdict
    from dateutil.relativedelta import relativedelta
    
    employee = await db.employees.find_one({"email": user["email"]}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    emp_id = employee["employee_id"]
    
    # Get all leave records
    leave_records = await db.leave_records.find(
        {"employee_id": emp_id},
        {"_id": 0}
    ).to_list(10000)
    
    # Calculate years based on joining date
    joining_date = datetime.fromisoformat(employee["joining_date"].replace('Z', '+00:00'))
    if joining_date.tzinfo is not None:
        joining_date = joining_date.replace(tzinfo=None)
    
    current_date = datetime.now()
    
    years_diff = relativedelta(current_date, joining_date)
    total_years = years_diff.years + (1 if years_diff.months > 0 or years_diff.days > 0 else 0)
    if total_years == 0:
        total_years = 1
    
    years_data = []
    
    for year_num in range(1, total_years + 1):
        year_start = joining_date + relativedelta(years=year_num - 1)
        year_end = joining_date + relativedelta(years=year_num) - relativedelta(days=1)
        encash_month = year_start + relativedelta(years=1, months=1)
        
        is_current_year = year_start <= current_date <= year_end
        is_closed = current_date > year_end
        
        year_leaves = []
        for l in leave_records:
            try:
                leave_date = datetime.fromisoformat(l["date"].replace('Z', '+00:00'))
                if leave_date.tzinfo is not None:
                    leave_date = leave_date.replace(tzinfo=None)
                if year_start <= leave_date <= year_end:
                    year_leaves.append(l)
            except:
                continue
        
        monthly_leaves = defaultdict(lambda: {"pl": [], "cl": [], "pl_total": 0, "cl_total": 0})
        
        for record in year_leaves:
            try:
                record_date = datetime.fromisoformat(record["date"].replace('Z', '+00:00'))
                if record_date.tzinfo is not None:
                    record_date = record_date.replace(tzinfo=None)
                month_key = record_date.strftime("%B %Y")
                
                leave_entry = {
                    "id": record["id"],
                    "date": record["date"],
                    "type": record["leave_type"],
                    "days": record["leave_days"],
                    "status": record["status"]
                }
                
                if record["leave_type"] == "PL":
                    monthly_leaves[month_key]["pl"].append(leave_entry)
                    monthly_leaves[month_key]["pl_total"] += record["leave_days"]
                else:
                    monthly_leaves[month_key]["cl"].append(leave_entry)
                    monthly_leaves[month_key]["cl_total"] += record["leave_days"]
            except:
                continue
        
        sorted_months = sorted(
            monthly_leaves.items(),
            key=lambda x: datetime.strptime(x[0], "%B %Y"),
            reverse=True
        )
        
        pl_taken = sum(r["leave_days"] for r in year_leaves if r["leave_type"] == "PL")
        cl_taken = sum(r["leave_days"] for r in year_leaves if r["leave_type"] == "CL")
        settled_pl = 16 - pl_taken if is_closed else 0
        available_pl = 16 - pl_taken if is_current_year else 0
        
        years_data.append({
            "year_number": year_num,
            "year_label": f"Year {year_num}",
            "start_date": year_start.strftime("%Y-%m-%d"),
            "end_date": year_end.strftime("%Y-%m-%d"),
            "encash_month": encash_month.strftime("%B %Y"),
            "is_current": is_current_year,
            "is_closed": is_closed,
            "pl_taken": pl_taken,
            "cl_taken": cl_taken,
            "settled_pl": settled_pl,
            "available_pl": available_pl,
            "monthly_leaves": [{"month": k, "leaves": v} for k, v in sorted_months]
        })
    
    current_year_data = next((y for y in years_data if y["is_current"]), years_data[-1] if years_data else None)
    
    return {
        "employee_id": employee["employee_id"],
        "name": employee["name"],
        "joining_date": employee["joining_date"],
        "total_years": total_years,
        "current_year_pl_taken": current_year_data["pl_taken"] if current_year_data else 0,
        "current_year_cl_taken": current_year_data["cl_taken"] if current_year_data else 0,
        "current_year_available_pl": current_year_data["available_pl"] if current_year_data else 16,
        "years_data": years_data
    }

@api_router.get("/leaves/employee-yearwise/{emp_id}")
async def get_employee_yearwise_leaves(emp_id: str, user: dict = Depends(get_current_user)):
    from collections import defaultdict
    from dateutil.relativedelta import relativedelta
    
    employee = await db.employees.find_one({"employee_id": emp_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Get all leave records
    leave_records = await db.leave_records.find(
        {"employee_id": emp_id},
        {"_id": 0}
    ).to_list(10000)
    
    # Get encashment history
    encashments = await db.leave_encashments.find(
        {"employee_id": emp_id},
        {"_id": 0}
    ).to_list(1000)
    
    # Calculate years based on joining date - use naive datetimes for consistency
    joining_date = datetime.fromisoformat(employee["joining_date"].replace('Z', '+00:00'))
    if joining_date.tzinfo is not None:
        joining_date = joining_date.replace(tzinfo=None)
    
    current_date = datetime.now()
    
    # Calculate how many complete years
    years_diff = relativedelta(current_date, joining_date)
    total_years = years_diff.years + (1 if years_diff.months > 0 or years_diff.days > 0 else 0)
    if total_years == 0:
        total_years = 1  # At least 1 year for new employees
    
    # Group leaves by year (based on joining date anniversary)
    years_data = []
    
    for year_num in range(1, total_years + 1):
        year_start = joining_date + relativedelta(years=year_num - 1)
        year_end = joining_date + relativedelta(years=year_num) - relativedelta(days=1)
        # Encashment month = Leave year completion + 1 month (13th month from year start)
        encash_month = year_start + relativedelta(years=1, months=1)
        
        is_current_year = year_start <= current_date <= year_end
        is_closed = current_date > year_end
        
        # Filter leaves for this year
        year_leaves = []
        for l in leave_records:
            try:
                leave_date = datetime.fromisoformat(l["date"].replace('Z', '+00:00'))
                if leave_date.tzinfo is not None:
                    leave_date = leave_date.replace(tzinfo=None)
                if year_start <= leave_date <= year_end:
                    year_leaves.append(l)
            except:
                continue
        
        # Group by month
        monthly_leaves = defaultdict(lambda: {"pl": [], "cl": [], "pl_total": 0, "cl_total": 0})
        
        for record in year_leaves:
            try:
                record_date = datetime.fromisoformat(record["date"].replace('Z', '+00:00'))
                if record_date.tzinfo is not None:
                    record_date = record_date.replace(tzinfo=None)
                month_key = record_date.strftime("%B %Y")
                
                leave_entry = {
                    "id": record["id"],
                    "date": record["date"],
                    "type": record["leave_type"],
                    "days": record["leave_days"],
                    "status": record["status"]
                }
                
                if record["leave_type"] == "PL":
                    monthly_leaves[month_key]["pl"].append(leave_entry)
                    monthly_leaves[month_key]["pl_total"] += record["leave_days"]
                else:
                    monthly_leaves[month_key]["cl"].append(leave_entry)
                    monthly_leaves[month_key]["cl_total"] += record["leave_days"]
            except:
                continue
        
        # Sort months
        sorted_months = sorted(
            monthly_leaves.items(),
            key=lambda x: datetime.strptime(x[0], "%B %Y"),
            reverse=True
        )
        
        # Calculate totals
        pl_taken = sum(r["leave_days"] for r in year_leaves if r["leave_type"] == "PL")
        cl_taken = sum(r["leave_days"] for r in year_leaves if r["leave_type"] == "CL")
        settled_pl = 16 - pl_taken if is_closed else 0
        available_pl = 16 - pl_taken if is_current_year else 0
        
        years_data.append({
            "year_number": year_num,
            "year_label": f"Year {year_num}",
            "start_date": year_start.strftime("%Y-%m-%d"),
            "end_date": year_end.strftime("%Y-%m-%d"),
            "encash_month": encash_month.strftime("%B %Y"),
            "is_current": is_current_year,
            "is_closed": is_closed,
            "pl_taken": pl_taken,
            "cl_taken": cl_taken,
            "settled_pl": settled_pl,
            "available_pl": available_pl,
            "monthly_leaves": [{"month": k, "leaves": v} for k, v in sorted_months]
        })
    
    # Current year data for main list
    current_year_data = next((y for y in years_data if y["is_current"]), years_data[-1] if years_data else None)
    
    return {
        "employee_id": employee["employee_id"],
        "name": employee["name"],
        "joining_date": employee["joining_date"],
        "total_years": total_years,
        "current_year_pl_taken": current_year_data["pl_taken"] if current_year_data else 0,
        "current_year_cl_taken": current_year_data["cl_taken"] if current_year_data else 0,
        "current_year_available_pl": current_year_data["available_pl"] if current_year_data else 16,
        "years_data": years_data
    }

@api_router.get("/leaves/tracker-yearwise")
async def get_employee_detailed_leaves(emp_id: str, user: dict = Depends(get_current_user)):
    employee = await db.employees.find_one({"employee_id": emp_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Get all leave records
    leave_records = await db.leave_records.find(
        {"employee_id": emp_id},
        {"_id": 0}
    ).to_list(1000)
    
    # Get encashment history
    encashments = await db.leave_encashments.find(
        {"employee_id": emp_id},
        {"_id": 0}
    ).to_list(1000)
    
    # Calculate current cycle
    joining_date = datetime.fromisoformat(employee["joining_date"])
    probation_end = datetime.fromisoformat(employee["probation_end_date"])
    if probation_end.tzinfo is None:
        probation_end = probation_end.replace(tzinfo=timezone.utc)
    is_probation = get_ist_now() < probation_end
    
    # Group leaves by month-year
    from collections import defaultdict
    monthly_leaves = defaultdict(lambda: {"pl": [], "cl": [], "pl_total": 0, "cl_total": 0})
    
    for record in leave_records:
        record_date = datetime.fromisoformat(record["date"])
        month_key = record_date.strftime("%B %Y")  # e.g., "January 2026"
        
        leave_entry = {
            "id": record["id"],
            "date": record["date"],
            "type": record["leave_type"],
            "days": record["leave_days"],
            "status": record["status"]
        }
        
        if record["leave_type"] == "PL":
            monthly_leaves[month_key]["pl"].append(leave_entry)
            monthly_leaves[month_key]["pl_total"] += record["leave_days"]
        else:
            monthly_leaves[month_key]["cl"].append(leave_entry)
            monthly_leaves[month_key]["cl_total"] += record["leave_days"]
    
    # Sort months chronologically
    sorted_months = sorted(
        monthly_leaves.items(),
        key=lambda x: datetime.strptime(x[0], "%B %Y"),
        reverse=True
    )
    
    # Calculate totals
    pl_taken = sum(r["leave_days"] for r in leave_records if r["leave_type"] == "PL")
    cl_taken = sum(r["leave_days"] for r in leave_records if r["leave_type"] == "CL")
    available_pl = 0 if is_probation else employee["annual_pl_allocation"] - pl_taken
    
    # Find last encashment month
    last_encashment = encashments[-1] if encashments else None
    
    return {
        "employee_id": employee["employee_id"],
        "name": employee["name"],
        "joining_date": employee["joining_date"],
        "is_probation": is_probation,
        "pl_allocated": 0 if is_probation else employee["annual_pl_allocation"],
        "pl_taken": pl_taken,
        "cl_taken": cl_taken,
        "available_pl": available_pl,
        "monthly_leaves": [{"month": k, "leaves": v} for k, v in sorted_months],
        "encashments": encashments,
        "last_encashment": last_encashment
    }

@api_router.put("/leaves/records/{record_id}")
async def update_leave_record(
    record_id: str,
    leave_type: str,
    leave_days: float,
    leave_date: Optional[str] = None,
    user: dict = Depends(require_role(["admin", "hr"]))
):
    record = await db.leave_records.find_one({"id": record_id})
    if not record:
        raise HTTPException(status_code=404, detail="Leave record not found")
    
    old_type = record["leave_type"]
    old_days = record["leave_days"]
    
    # Build update data
    update_data = {
        "leave_type": leave_type,
        "leave_days": leave_days,
        "updated_at": get_ist_now_iso()
    }
    
    # Update date if provided
    if leave_date:
        update_data["date"] = leave_date
    
    # Update the record
    await db.leave_records.update_one(
        {"id": record_id},
        {"$set": update_data}
    )
    
    # Update employee totals
    employee = await db.employees.find_one({"employee_id": record["employee_id"]})
    if employee:
        # Reverse old counts
        if old_type == "PL":
            await db.employees.update_one(
                {"employee_id": record["employee_id"]},
                {"$inc": {"pl_taken": -old_days}}
            )
        else:
            await db.employees.update_one(
                {"employee_id": record["employee_id"]},
                {"$inc": {"cl_taken": -old_days}}
            )
        
        # Add new counts
        if leave_type == "PL":
            await db.employees.update_one(
                {"employee_id": record["employee_id"]},
                {"$inc": {"pl_taken": leave_days}}
            )
        else:
            await db.employees.update_one(
                {"employee_id": record["employee_id"]},
                {"$inc": {"cl_taken": leave_days}}
            )
    
    return {"message": "Leave record updated"}

@api_router.delete("/leaves/records/{record_id}")
async def delete_leave_record(record_id: str, user: dict = Depends(require_role(["admin", "hr"]))):
    record = await db.leave_records.find_one({"id": record_id})
    if not record:
        raise HTTPException(status_code=404, detail="Leave record not found")
    
    # Update employee totals
    if record["leave_type"] == "PL":
        await db.employees.update_one(
            {"employee_id": record["employee_id"]},
            {"$inc": {"pl_taken": -record["leave_days"]}}
        )
    else:
        await db.employees.update_one(
            {"employee_id": record["employee_id"]},
            {"$inc": {"cl_taken": -record["leave_days"]}}
        )
    
    await db.leave_records.delete_one({"id": record_id})
    return {"message": "Leave record deleted"}

@api_router.put("/leaves/applications/{leave_id}/approve")
async def approve_leave(leave_id: str, approval: LeaveApproval, user: dict = Depends(require_role(["admin", "hr"]))):
    leave_app = await db.leave_applications.find_one({"id": leave_id})
    if not leave_app:
        raise HTTPException(status_code=404, detail="Leave application not found")
    
    now = get_ist_now_iso()
    
    # Determine final status based on leave_dates
    # All approved → "approved", All rejected → "rejected"
    # NO partial status - we just use approved/rejected based on what dates are included
    final_status = approval.status
    if approval.leave_dates:
        approved_count = sum(1 for ld in approval.leave_dates if ld.leave_type.lower() != "rejected")
        rejected_count = sum(1 for ld in approval.leave_dates if ld.leave_type.lower() == "rejected")
        
        if approved_count == 0 and rejected_count > 0:
            final_status = "rejected"  # All rejected
        else:
            final_status = "approved"  # Has at least some approved dates
    
    # Update application with leave_dates if provided
    update_data = {
        "status": final_status,
        "approved_by": user["username"],
        "approved_date": now,
        "comments": approval.comments,
        "reject_reason": approval.reject_reason if final_status == "rejected" else ""
    }
    
    if approval.leave_dates:
        update_data["leave_dates"] = [ld.dict() for ld in approval.leave_dates]
    
    await db.leave_applications.update_one(
        {"id": leave_id},
        {"$set": update_data}
    )
    
    # If has approved dates, create leave records for approved dates only
    if final_status == "approved":
        total_pl = 0.0
        total_cl = 0.0
        
        # Use leave_dates if provided, otherwise fall back to old logic
        if approval.leave_dates:
            for leave_date in approval.leave_dates:
                # Skip rejected dates
                if leave_date.leave_type.lower() == "rejected":
                    continue
                    
                # Handle PL/2 & CL/2 special case - creates two records (0.5 PL + 0.5 CL)
                if leave_date.leave_type == "PL/2 & CL/2":
                    # Create PL record (0.5)
                    pl_record_id = str(uuid.uuid4())
                    pl_record = {
                        "id": pl_record_id,
                        "employee_id": leave_app["employee_id"],
                        "date": leave_date.date,
                        "leave_type": "PL",
                        "leave_days": 0.5,
                        "status": "taken",
                        "approved_by": user["username"],
                        "applied_date": leave_app["created_at"],
                        "approved_date": now,
                        "created_at": now,
                        "updated_at": now
                    }
                    await db.leave_records.insert_one(pl_record)
                    total_pl += 0.5
                    
                    # Create CL record (0.5)
                    cl_record_id = str(uuid.uuid4())
                    cl_record = {
                        "id": cl_record_id,
                        "employee_id": leave_app["employee_id"],
                        "date": leave_date.date,
                        "leave_type": "CL",
                        "leave_days": 0.5,
                        "status": "taken",
                        "approved_by": user["username"],
                        "applied_date": leave_app["created_at"],
                        "approved_date": now,
                        "created_at": now,
                        "updated_at": now
                    }
                    await db.leave_records.insert_one(cl_record)
                    total_cl += 0.5
                else:
                    # Standard leave types: PL, CL, Half PL, Half CL
                    record_id = str(uuid.uuid4())
                    
                    # Determine leave days based on type
                    leave_days = 0.5 if 'Half' in leave_date.leave_type else 1.0
                    leave_type_clean = 'PL' if 'PL' in leave_date.leave_type else 'CL'
                    
                    leave_record = {
                        "id": record_id,
                        "employee_id": leave_app["employee_id"],
                        "date": leave_date.date,
                        "leave_type": leave_type_clean,
                        "leave_days": leave_days,
                        "status": "taken",
                        "approved_by": user["username"],
                        "applied_date": leave_app["created_at"],
                        "approved_date": now,
                        "created_at": now,
                        "updated_at": now
                    }
                    await db.leave_records.insert_one(leave_record)
                    
                    if leave_type_clean == "PL":
                        total_pl += leave_days
                    else:
                        total_cl += leave_days
        else:
            # Fall back to old logic if no leave_dates provided
            from_date = datetime.fromisoformat(leave_app["from_date"])
            to_date = datetime.fromisoformat(leave_app["to_date"])
            current_date = from_date
            default_type = leave_app.get("leave_type", "PL") or "PL"
            
            while current_date <= to_date:
                record_id = str(uuid.uuid4())
                leave_record = {
                    "id": record_id,
                    "employee_id": leave_app["employee_id"],
                    "date": current_date.isoformat(),
                    "leave_type": default_type,
                    "leave_days": 1.0,
                    "status": "taken",
                    "approved_by": user["username"],
                    "applied_date": leave_app["created_at"],
                    "approved_date": now,
                    "created_at": now,
                    "updated_at": now
                }
                await db.leave_records.insert_one(leave_record)
                
                if default_type == "PL":
                    total_pl += 1.0
                else:
                    total_cl += 1.0
                    
                current_date += timedelta(days=1)
        
        # Update employee leave counts
        if total_pl > 0:
            await db.employees.update_one(
                {"employee_id": leave_app["employee_id"]},
                {"$inc": {"pl_taken": total_pl}}
            )
        if total_cl > 0:
            await db.employees.update_one(
                {"employee_id": leave_app["employee_id"]},
                {"$inc": {"cl_taken": total_cl}}
            )
    
    # === EMAIL TRIGGER (Post-action - does NOT affect leave logic above) ===
    try:
        # Get employee details for email
        employee = await db.employees.find_one({"employee_id": leave_app["employee_id"]})
        if employee and employee.get("email"):
            # Prepare applied dates
            from_date = datetime.fromisoformat(leave_app["from_date"])
            to_date = datetime.fromisoformat(leave_app["to_date"])
            applied_dates = []
            current = from_date
            while current <= to_date:
                applied_dates.append(current.strftime("%d %b %Y"))
                current += timedelta(days=1)
            
            # Prepare approved and rejected dates
            approved_dates = []
            rejected_dates = []
            
            if approval.leave_dates:
                for ld in approval.leave_dates:
                    date_str = datetime.fromisoformat(ld.date).strftime("%d %b %Y")
                    if ld.leave_type.lower() == "rejected":
                        rejected_dates.append({
                            "date": date_str,
                            "reason": ld.reject_reason or approval.reject_reason or "Not specified"
                        })
                    else:
                        approved_dates.append({
                            "date": date_str,
                            "type": ld.leave_type
                        })
            elif final_status == "approved":
                # All dates approved with same type
                for d in applied_dates:
                    approved_dates.append({
                        "date": d,
                        "type": leave_app.get("leave_type", "PL") or "PL"
                    })
            elif final_status == "rejected":
                # All dates rejected
                for d in applied_dates:
                    rejected_dates.append({
                        "date": d,
                        "reason": approval.reject_reason or "Not specified"
                    })
            
            # Send email notification (async, non-blocking)
            await send_leave_notification_email(
                employee_email=employee["email"],
                employee_name=employee["name"],
                leave_type=leave_app.get("leave_type", ""),
                applied_dates=applied_dates,
                approved_dates=approved_dates,
                rejected_dates=rejected_dates,
                status=final_status,
                approved_by=user["username"],
                comments=approval.comments or ""
            )
    except Exception as e:
        # Email failure should NOT affect leave approval response
        logging.error(f"Failed to send leave notification email: {str(e)}")
    # === END EMAIL TRIGGER ===
    
    return {"message": "Leave application processed"}

@api_router.post("/leaves/import")
async def import_leaves(file: UploadFile = File(...), user: dict = Depends(require_role(["admin", "hr"]))):
    contents = await file.read()
    decoded = contents.decode('utf-8')
    csv_reader = csv.DictReader(io.StringIO(decoded))
    
    imported = 0
    errors = []
    now = get_ist_now_iso()
    
    for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 because row 1 is header
        try:
            # Validate required fields
            if not row.get("employee_id"):
                errors.append(f"Row {row_num}: Missing employee_id")
                continue
            if not row.get("date"):
                errors.append(f"Row {row_num}: Missing date")
                continue
            if not row.get("type"):
                errors.append(f"Row {row_num}: Missing type (PL/CL)")
                continue
            if not row.get("leave"):
                errors.append(f"Row {row_num}: Missing leave days")
                continue
            
            # Validate date format (YYYY-MM-DD)
            date_str = row["date"].strip()
            try:
                datetime.fromisoformat(date_str)
            except ValueError:
                errors.append(f"Row {row_num}: Invalid date format '{date_str}'. Use YYYY-MM-DD")
                continue
            
            # Validate leave type
            leave_type = row["type"].strip().upper()
            if leave_type not in ["PL", "CL"]:
                errors.append(f"Row {row_num}: Invalid leave type '{row['type']}'. Use PL or CL")
                continue
            
            # Validate leave days
            try:
                leave_days = float(row["leave"])
                if leave_days <= 0:
                    errors.append(f"Row {row_num}: Leave days must be positive")
                    continue
            except ValueError:
                errors.append(f"Row {row_num}: Invalid leave days '{row['leave']}'")
                continue
            
            # Validate employee exists
            employee = await db.employees.find_one({"employee_id": row["employee_id"].strip()})
            if not employee:
                errors.append(f"Row {row_num}: Employee ID '{row['employee_id']}' not found")
                continue
            
            record_id = str(uuid.uuid4())
            leave_record = {
                "id": record_id,
                "employee_id": row["employee_id"].strip(),
                "date": date_str,
                "leave_type": leave_type,
                "leave_days": leave_days,
                "status": row.get("status", "taken").strip() or "taken",
                "approved_by": user["username"],
                "applied_date": now,
                "approved_date": now,
                "created_at": now,
                "updated_at": now
            }
            await db.leave_records.insert_one(leave_record)
            
            # Update employee leave count
            if leave_type == "PL":
                await db.employees.update_one(
                    {"employee_id": row["employee_id"].strip()},
                    {"$inc": {"pl_taken": leave_days}}
                )
            else:
                await db.employees.update_one(
                    {"employee_id": row["employee_id"].strip()},
                    {"$inc": {"cl_taken": leave_days}}
                )
            
            imported += 1
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
    
    if errors and imported > 0:
        return {"message": f"Imported {imported} leave records with {len(errors)} errors", "errors": errors[:10]}
    elif errors:
        raise HTTPException(status_code=400, detail=f"Import failed. Errors: {'; '.join(errors[:5])}")
    
    return {"message": f"Successfully imported {imported} leave records"}

class EncashmentRequest(BaseModel):
    employee_id: str
    encash_month: str

@api_router.post("/leaves/encash")
async def encash_leave(request: EncashmentRequest, user: dict = Depends(require_role(["admin", "hr"]))):
    employee = await db.employees.find_one({"employee_id": request.employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Get all leave records for this employee
    leave_records = await db.leave_records.find(
        {"employee_id": request.employee_id},
        {"_id": 0}
    ).to_list(1000)
    
    pl_taken = sum(r["leave_days"] for r in leave_records if r["leave_type"] == "PL")
    pl_remaining = employee["annual_pl_allocation"] - pl_taken
    
    if pl_remaining <= 0:
        raise HTTPException(status_code=400, detail="No remaining PL to encash")
    
    # Create encashment record
    encash_id = str(uuid.uuid4())
    now = get_ist_now_iso()
    encashment_record = {
        "id": encash_id,
        "employee_id": request.employee_id,
        "encash_month": request.encash_month,
        "pl_encashed": pl_remaining,
        "processed_by": user["username"],
        "processed_date": now,
        "created_at": now
    }
    
    await db.leave_encashments.insert_one(encashment_record)
    
    # Reset PL taken for new cycle
    await db.employees.update_one(
        {"employee_id": request.employee_id},
        {"$set": {"pl_taken": 0.0, "cl_taken": 0.0}}
    )
    
    return {
        "message": f"Encashed {pl_remaining} PL for employee {request.employee_id}",
        "pl_encashed": pl_remaining,
        "encash_month": request.encash_month
    }

@api_router.get("/leaves/encashments")
async def get_encashments(user: dict = Depends(require_role(["admin", "hr"]))):
    encashments = await db.leave_encashments.find({}, {"_id": 0}).to_list(1000)
    return encashments


@api_router.get("/leaves/tracker")
async def get_leave_tracker(
    employee_status: str = Query(default="active", description="Filter by employee status: active, ex-employee, all"),
    user: dict = Depends(require_role(["admin", "hr"]))
):
    from dateutil.relativedelta import relativedelta
    from collections import defaultdict
    
    # Filter employees by status
    if employee_status and employee_status != "all":
        employees = await db.employees.find({"status": employee_status}, {"_id": 0}).to_list(1000)
    else:
        employees = await db.employees.find({}, {"_id": 0}).to_list(1000)
    
    current_date = get_ist_now()
    
    # Fetch ALL leave records in a single query (fixes N+1 query issue)
    all_leave_records = await db.leave_records.find({}, {"_id": 0}).to_list(100000)
    
    # Group leave records by employee_id in memory
    leave_by_employee = defaultdict(list)
    for record in all_leave_records:
        leave_by_employee[record["employee_id"]].append(record)
    
    tracker = []
    for emp in employees:
        joining_date = datetime.fromisoformat(emp["joining_date"])
        if joining_date.tzinfo is None:
            joining_date = joining_date.replace(tzinfo=timezone.utc)
        
        # Calculate current leave year based on joining date
        years_completed = relativedelta(current_date, joining_date).years
        current_year_start = joining_date + relativedelta(years=years_completed)
        current_year_end = joining_date + relativedelta(years=years_completed + 1) - relativedelta(days=1)
        
        # Get leave records for this employee from memory (no DB query)
        leave_records = leave_by_employee.get(emp["employee_id"], [])
        
        # Filter leaves for current year only (with error handling for bad dates)
        current_year_leaves = []
        for r in leave_records:
            try:
                record_date = datetime.fromisoformat(r["date"]).replace(tzinfo=timezone.utc)
                if current_year_start <= record_date <= current_year_end:
                    current_year_leaves.append(r)
            except (ValueError, TypeError) as e:
                # Skip records with invalid dates
                logging.warning(f"Invalid date in leave record {r.get('id')}: {r.get('date')} - {e}")
                continue
        
        pl_taken = sum(r.get("leave_days", 1.0) for r in current_year_leaves if r.get("leave_type") == "PL")
        cl_taken = sum(r.get("leave_days", 1.0) for r in current_year_leaves if r.get("leave_type") == "CL")
        available_pl = 16 - pl_taken
        
        tracker.append({
            "employee_id": emp["employee_id"],
            "name": emp["name"],
            "joining_date": emp["joining_date"],
            "status": emp.get("status", "active"),
            "pl_taken": pl_taken,
            "cl_taken": cl_taken,
            "available_pl": available_pl
        })
    
    return tracker

@api_router.delete("/leaves/reset/{emp_id}")
async def reset_employee_leaves(emp_id: str, user: dict = Depends(require_role(["admin"]))):
    """Reset all leave records for a specific employee (Admin only)"""
    # Verify employee exists
    employee = await db.employees.find_one({"employee_id": emp_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Delete all leave records for this employee
    delete_result = await db.leave_records.delete_many({"employee_id": emp_id})
    
    # Reset pl_taken and cl_taken counters
    await db.employees.update_one(
        {"employee_id": emp_id},
        {"$set": {"pl_taken": 0, "cl_taken": 0}}
    )
    
    return {
        "message": f"Successfully reset leave data for {employee.get('name', emp_id)}",
        "deleted_records": delete_result.deleted_count
    }

@api_router.get("/leaves/employee/{emp_id}")
async def get_employee_leaves(emp_id: str, user: dict = Depends(get_current_user)):
    employee = await db.employees.find_one({"employee_id": emp_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    leave_records = await db.leave_records.find(
        {"employee_id": emp_id},
        {"_id": 0}
    ).to_list(1000)
    
    pl_taken = sum(r["leave_days"] for r in leave_records if r["leave_type"] == "PL")
    cl_taken = sum(r["leave_days"] for r in leave_records if r["leave_type"] == "CL")
    
    joining_date = datetime.fromisoformat(employee["joining_date"])
    probation_end = datetime.fromisoformat(employee["probation_end_date"])
    # Ensure both datetimes are timezone-aware for comparison
    if probation_end.tzinfo is None:
        probation_end = probation_end.replace(tzinfo=timezone.utc)
    is_probation = get_ist_now() < probation_end
    
    return {
        "employee_id": employee["employee_id"],
        "name": employee["name"],
        "joining_date": employee["joining_date"],
        "is_probation": is_probation,
        "pl_allocated": 0 if is_probation else employee["annual_pl_allocation"],
        "pl_taken": pl_taken,
        "cl_taken": cl_taken,
        "pl_remaining": 0 if is_probation else employee["annual_pl_allocation"] - pl_taken,
        "leave_records": leave_records
    }

# Dashboard Routes
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(user: dict = Depends(get_current_user)):
    total_employees = await db.employees.count_documents({"status": "active"})
    total_departments = await db.departments.count_documents({"is_active": True})
    total_projects = await db.projects.count_documents({})
    active_projects = await db.projects.count_documents({"status": {"$in": ["active", "running"]}})
    
    return {
        "total_employees": total_employees,
        "total_departments": total_departments,
        "total_projects": total_projects,
        "active_projects": active_projects
    }

@api_router.get("/dashboard/employees-on-leave")
async def get_employees_on_leave(user: dict = Depends(require_role(["admin", "hr"]))):
    """Get employees on leave today and tomorrow"""
    
    # Get today and tomorrow in IST
    ist = ZoneInfo('Asia/Kolkata')
    now = datetime.now(ist)
    today = now.strftime('%Y-%m-%d')
    tomorrow = (now + timedelta(days=1)).strftime('%Y-%m-%d')
    
    # Get all approved leave applications
    approved_leaves = await db.leave_applications.find({"status": "approved"}, {"_id": 0}).to_list(10000)
    
    # Get employee info
    employees = await db.employees.find({"status": "active"}, {"_id": 0}).to_list(1000)
    emp_map = {emp["employee_id"]: emp for emp in employees}
    
    today_leaves = []
    tomorrow_leaves = []
    
    for leave in approved_leaves:
        from_date = leave.get("from_date", "")
        to_date = leave.get("to_date", "")
        emp_id = leave.get("employee_id", "")
        emp = emp_map.get(emp_id, {})
        
        if not emp:
            continue
        
        # Check leave_dates_input for half-day info
        leave_dates_input = leave.get("leave_dates_input", [])
        leave_dates = leave.get("leave_dates", [])  # From approval
        
        # Check if today falls within the leave range
        if from_date <= today <= to_date:
            day_type = "full"
            # Check for half-day
            if leave_dates_input:
                for ld in leave_dates_input:
                    if ld.get("date") == today and ld.get("day_type") == "half":
                        day_type = "half"
                        break
            if leave_dates:
                for ld in leave_dates:
                    if ld.get("date") == today and "Half" in ld.get("leave_type", ""):
                        day_type = "half"
                        break
            
            today_leaves.append({
                "employee_id": emp_id,
                "name": emp.get("name", "Unknown"),
                "day_type": day_type
            })
        
        # Check if tomorrow falls within the leave range
        if from_date <= tomorrow <= to_date:
            day_type = "full"
            # Check for half-day
            if leave_dates_input:
                for ld in leave_dates_input:
                    if ld.get("date") == tomorrow and ld.get("day_type") == "half":
                        day_type = "half"
                        break
            if leave_dates:
                for ld in leave_dates:
                    if ld.get("date") == tomorrow and "Half" in ld.get("leave_type", ""):
                        day_type = "half"
                        break
            
            tomorrow_leaves.append({
                "employee_id": emp_id,
                "name": emp.get("name", "Unknown"),
                "day_type": day_type
            })
    
    return {
        "today": today_leaves,
        "tomorrow": tomorrow_leaves
    }

@api_router.get("/dashboard/admin-analytics")
async def get_admin_dashboard_analytics(user: dict = Depends(require_role(["admin", "hr"]))):
    """Comprehensive analytics for Admin/HR dashboard"""
    from collections import defaultdict
    
    current_date = datetime.now()
    
    # Get all projects
    projects = await db.projects.find({}, {"_id": 0}).to_list(10000)
    
    # Project Status Overview
    total_projects = len(projects)
    completed_projects = len([p for p in projects if p.get("status") == "completed"])
    in_progress_projects = len([p for p in projects if p.get("status") in ["active", "running", "in-progress", "ongoing"]])
    
    # Calculate late projects:
    # ONLY count projects where status is explicitly "late"
    late_projects = []
    on_time_projects = []
    
    for project in projects:
        # Only count as late if status is explicitly "late"
        if project.get("status") == "late":
            late_projects.append(project)
            continue
            
        # For completed projects, check if they were completed on time
        if project.get("status") == "completed":
            end_date_str = project.get("end_date")
            if end_date_str:
                try:
                    end_date = datetime.fromisoformat(end_date_str.replace('Z', '+00:00'))
                    if end_date.tzinfo:
                        end_date = end_date.replace(tzinfo=None)
                    
                    completed_date_str = project.get("updated_at", project.get("created_at"))
                    if completed_date_str:
                        completed_date = datetime.fromisoformat(completed_date_str.replace('Z', '+00:00'))
                        if completed_date.tzinfo:
                            completed_date = completed_date.replace(tzinfo=None)
                        if completed_date <= end_date:
                            on_time_projects.append(project)
                    else:
                        on_time_projects.append(project)
                except:
                    on_time_projects.append(project)
            else:
                on_time_projects.append(project)
    
    # Get all employees (including inactive for historical tracking)
    employees = await db.employees.find({}, {"_id": 0}).to_list(1000)
    emp_map = {emp["employee_id"]: emp["name"] for emp in employees}
    
    # Get only active employees for performance tracking
    active_employees = {emp["employee_id"] for emp in employees if emp.get("status") == "active"}
    
    # Employee-wise project analysis - track ALL assigned employees even if not in our system
    employee_performance = defaultdict(lambda: {
        "name": "",
        "employee_id": "",
        "total_assigned": 0,
        "completed_on_time": 0,
        "completed_late": 0,
        "ongoing": 0
    })
    
    for project in projects:
        assigned_employees = project.get("assigned_employees", [])
        
        for emp_id in assigned_employees:
            # Use employee name if found, otherwise use the ID as name
            emp_name = emp_map.get(emp_id, f"Employee {emp_id}")
            employee_performance[emp_id]["name"] = emp_name
            employee_performance[emp_id]["employee_id"] = emp_id
            employee_performance[emp_id]["total_assigned"] += 1
            
            if project.get("status") == "completed":
                # Check if this project was completed late
                if project in late_projects:
                    employee_performance[emp_id]["completed_late"] += 1
                else:
                    employee_performance[emp_id]["completed_on_time"] += 1
            elif project.get("status") == "late" or project in late_projects:
                employee_performance[emp_id]["completed_late"] += 1
            else:
                employee_performance[emp_id]["ongoing"] += 1
    
    # Top performers (least late) and Low performers (most late) - ONLY ACTIVE EMPLOYEES
    performance_list = list(employee_performance.values())
    # Filter to only include active employees
    active_performance_list = [p for p in performance_list if p["employee_id"] in active_employees]
    active_performance_list.sort(key=lambda x: (-x["completed_on_time"], x["completed_late"]))
    
    top_performers = [p for p in active_performance_list if p["total_assigned"] > 0][:5]
    low_performers = sorted([p for p in active_performance_list if p["completed_late"] > 0], 
                           key=lambda x: x["completed_late"], reverse=True)[:5]
    
    # Monthly project completion trends (last 6 months)
    monthly_trends = []
    for i in range(5, -1, -1):
        month_start = current_date.replace(day=1) - timedelta(days=i*30)
        month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        month_name = month_start.strftime("%b %Y")
        
        completed_count = 0
        delayed_count = 0
        
        for project in projects:
            if project.get("status") == "completed":
                updated_at = project.get("updated_at", "")
                if updated_at:
                    try:
                        updated_date = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
                        if updated_date.tzinfo:
                            updated_date = updated_date.replace(tzinfo=None)
                        if month_start <= updated_date <= month_end:
                            completed_count += 1
                            if project in late_projects:
                                delayed_count += 1
                    except:
                        continue
        
        monthly_trends.append({
            "month": month_name,
            "completed": completed_count,
            "delayed": delayed_count
        })
    
    # Late projects by employee - include all assigned employees even if not in system
    late_by_employee = defaultdict(lambda: {"name": "", "count": 0, "projects": []})
    for project in late_projects:
        for emp_id in project.get("assigned_employees", []):
            emp_name = emp_map.get(emp_id, f"Employee {emp_id}")
            late_by_employee[emp_id]["name"] = emp_name
            late_by_employee[emp_id]["count"] += 1
            late_by_employee[emp_id]["projects"].append(project.get("name", ""))
    
    # Active projects = ongoing + late + yet to start + NULL (not completed, hold, cancelled)
    active_projects = len([p for p in projects if p.get("status") not in ["completed", "hold", "cancelled", "cancel"]])
    
    return {
        "project_overview": {
            "total": total_projects,
            "active": active_projects,
            "completed": completed_projects,
            "in_progress": in_progress_projects,
            "late": len(late_projects),
            "on_time": len(on_time_projects)
        },
        "late_projects": [
            {
                "id": p.get("id"),
                "name": p.get("name"),
                "project_code": p.get("project_code"),
                "end_date": p.get("end_date"),
                "status": p.get("status"),
                "assigned_employees": p.get("assigned_employees", [])
            } for p in late_projects[:10]
        ],
        "late_by_employee": [
            {"employee_id": k, "name": v["name"], "count": v["count"], "projects": v["projects"][:3]}
            for k, v in sorted(late_by_employee.items(), key=lambda x: x[1]["count"], reverse=True)
            if k in active_employees  # Only include active employees
        ][:10],
        "employee_performance": active_performance_list,  # Only active employees
        "top_performers": top_performers,
        "low_performers": low_performers,
        "monthly_trends": monthly_trends
    }

@api_router.get("/dashboard/employee-analytics")
async def get_employee_dashboard_analytics(user: dict = Depends(get_current_user)):
    """Analytics for Employee dashboard - own data only"""
    from collections import defaultdict
    
    # Get employee details
    employee = await db.employees.find_one({"email": user["email"]}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    emp_id = employee["employee_id"]
    current_date = datetime.now()
    
    # Get projects assigned to this employee
    projects = await db.projects.find(
        {"assigned_employees": emp_id},
        {"_id": 0}
    ).to_list(1000)
    
    total_projects = len(projects)
    completed_projects = 0
    late_projects = 0
    on_time_projects = 0
    ongoing_projects = 0
    
    for project in projects:
        end_date_str = project.get("end_date")
        project_status = project.get("status", "").lower()
        
        # Check if project is explicitly marked as "late"
        if project_status == "late":
            late_projects += 1
            # Don't count late in ongoing - it's a separate status
            continue
        
        if project_status == "completed":
            completed_projects += 1
            if end_date_str:
                try:
                    end_date = datetime.fromisoformat(end_date_str.replace('Z', '+00:00'))
                    if end_date.tzinfo:
                        end_date = end_date.replace(tzinfo=None)
                    
                    updated_at = project.get("updated_at", "")
                    if updated_at:
                        completed_date = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
                        if completed_date.tzinfo:
                            completed_date = completed_date.replace(tzinfo=None)
                        if completed_date <= end_date:
                            on_time_projects += 1
                        # Don't add to late_projects here - already counted as completed
                except:
                    on_time_projects += 1
            else:
                on_time_projects += 1
        else:
            # Ongoing/In-progress/Active projects
            ongoing_projects += 1
            # Only mark as late if end_date has passed AND project is not completed
            # This should sync with admin's view - only admin-marked "late" status counts as late
            # Don't auto-calculate late based on end_date for consistency with admin view
    
    # Monthly personal trends (last 6 months)
    monthly_trends = []
    for i in range(5, -1, -1):
        month_start = current_date.replace(day=1) - timedelta(days=i*30)
        month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        month_name = month_start.strftime("%b %Y")
        
        completed_count = 0
        late_count = 0
        
        for project in projects:
            project_status = project.get("status", "")
            if project_status == "completed":
                updated_at = project.get("updated_at", "")
                end_date_str = project.get("end_date", "")
                if updated_at:
                    try:
                        updated_date = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
                        if updated_date.tzinfo:
                            updated_date = updated_date.replace(tzinfo=None)
                        if month_start <= updated_date <= month_end:
                            completed_count += 1
                            if end_date_str:
                                end_date = datetime.fromisoformat(end_date_str.replace('Z', '+00:00'))
                                if end_date.tzinfo:
                                    end_date = end_date.replace(tzinfo=None)
                                if updated_date > end_date:
                                    late_count += 1
                    except:
                        continue
        
        monthly_trends.append({
            "month": month_name,
            "completed": completed_count,
            "late": late_count,
            "on_time": max(0, completed_count - late_count)
        })
    
    # Get work hours summary
    work_entries = await db.work_entries.find(
        {"employee_id": emp_id},
        {"_id": 0}
    ).to_list(10000)
    
    total_hours = sum(e.get("hours", 0) for e in work_entries)
    
    # Hours by project
    hours_by_project = defaultdict(float)
    for entry in work_entries:
        hours_by_project[entry.get("project_id", "")] += entry.get("hours", 0)
    
    # Get project names
    project_hours = []
    for proj_id, hours in sorted(hours_by_project.items(), key=lambda x: x[1], reverse=True)[:5]:
        proj = next((p for p in projects if p.get("id") == proj_id), None)
        if proj:
            project_hours.append({
                "project_name": proj.get("name", "Unknown"),
                "hours": round(hours, 1)
            })
    
    # Leave summary
    leave_data = None
    try:
        # Get leave details
        leave_records = await db.leave_records.find(
            {"employee_id": emp_id},
            {"_id": 0}
        ).to_list(1000)
        
        joining_date = datetime.fromisoformat(employee["joining_date"].replace('Z', '+00:00'))
        if joining_date.tzinfo:
            joining_date = joining_date.replace(tzinfo=None)
        
        from dateutil.relativedelta import relativedelta
        years_diff = relativedelta(current_date, joining_date)
        year_start = joining_date + relativedelta(years=years_diff.years)
        year_end = year_start + relativedelta(years=1) - relativedelta(days=1)
        
        current_year_leaves = [
            r for r in leave_records
            if year_start <= datetime.fromisoformat(r["date"].replace('Z', '+00:00')).replace(tzinfo=None) <= year_end
        ]
        
        pl_taken = sum(r["leave_days"] for r in current_year_leaves if r["leave_type"] == "PL")
        cl_taken = sum(r["leave_days"] for r in current_year_leaves if r["leave_type"] == "CL")
        
        leave_data = {
            "available_pl": 16 - pl_taken,
            "pl_taken": pl_taken,
            "cl_taken": cl_taken
        }
    except:
        leave_data = {"available_pl": 16, "pl_taken": 0, "cl_taken": 0}
    
    return {
        "employee": {
            "name": employee["name"],
            "employee_id": emp_id,
            "department_ids": employee.get("department_ids", []),
            "joining_date": employee["joining_date"]
        },
        "project_summary": {
            "total": total_projects,
            "completed": completed_projects,
            "ongoing": ongoing_projects,
            "late": late_projects,
            "on_time": on_time_projects
        },
        "monthly_trends": monthly_trends,
        "total_work_hours": round(total_hours, 1),
        "hours_by_project": project_hours,
        "leave_summary": leave_data,
        "recent_projects": [
            {
                "name": p.get("name"),
                "status": p.get("status"),
                "end_date": p.get("end_date")
            } for p in projects[:5]
        ]
    }

# ==================== HOLIDAY MANAGEMENT ====================

@api_router.get("/holidays")
async def get_holidays(user: dict = Depends(get_current_user)):
    """Get all holidays - visible to all users"""
    holidays = []
    async for h in db.holidays.find({}, {"_id": 0}).sort("date", 1):
        holidays.append(h)
    return holidays

@api_router.get("/holidays/upcoming")
async def get_upcoming_holidays(user: dict = Depends(get_current_user)):
    """Get all upcoming holidays for the rest of the year"""
    today = get_ist_now().strftime("%Y-%m-%d")
    year_end = f"{get_ist_now().year}-12-31"
    
    holidays = []
    async for h in db.holidays.find(
        {"date": {"$gte": today, "$lte": year_end}},
        {"_id": 0}
    ).sort("date", 1):
        holidays.append(h)
    return holidays

@api_router.post("/holidays")
async def create_holiday(holiday: HolidayCreate, user: dict = Depends(require_role(["admin", "hr"]))):
    """Create a new holiday - Admin/HR only"""
    holiday_doc = {
        "id": str(uuid.uuid4()),
        "name": holiday.name,
        "date": holiday.date,
        "description": holiday.description or "",
        "created_at": get_ist_now_iso()
    }
    await db.holidays.insert_one(holiday_doc)
    return {k: v for k, v in holiday_doc.items() if k != "_id"}

@api_router.delete("/holidays/{holiday_id}")
async def delete_holiday(holiday_id: str, user: dict = Depends(require_role(["admin", "hr"]))):
    """Delete a holiday - Admin/HR only"""
    result = await db.holidays.delete_one({"id": holiday_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Holiday not found")
    return {"message": "Holiday deleted successfully"}

# ==================== UPCOMING BIRTHDAYS ====================

@api_router.get("/birthdays/upcoming")
async def get_upcoming_birthdays(user: dict = Depends(get_current_user)):
    """Get upcoming birthdays (next 30 days) based on employee birth_date"""
    today = get_ist_now()
    upcoming = []
    
    # Get all active employees with birth dates
    async for emp in db.employees.find(
        {"status": "active", "birth_date": {"$ne": None, "$exists": True}},
        {"_id": 0, "id": 1, "employee_id": 1, "name": 1, "birth_date": 1}
    ):
        if not emp.get("birth_date"):
            continue
        
        try:
            # Parse birth date
            birth_date = datetime.fromisoformat(emp["birth_date"].replace("Z", "+00:00"))
            
            # Calculate this year's birthday
            this_year_birthday = birth_date.replace(year=today.year)
            
            # If birthday already passed this year, check next year
            if this_year_birthday.date() < today.date():
                this_year_birthday = birth_date.replace(year=today.year + 1)
            
            # Check if within next 30 days
            days_until = (this_year_birthday.date() - today.date()).days
            
            if 0 <= days_until <= 30:
                upcoming.append({
                    "employee_id": emp["employee_id"],
                    "name": emp["name"],
                    "birth_date": emp["birth_date"],
                    "upcoming_date": this_year_birthday.strftime("%Y-%m-%d"),
                    "days_until": days_until
                })
        except Exception:
            continue
    
    # Sort by days until birthday
    upcoming.sort(key=lambda x: x["days_until"])
    return upcoming

# Middleware added here, router included at end of file
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

# ==================== BANKS MANAGEMENT ====================
# Banks CRUD - Admin Only

class BankCreate(BaseModel):
    name: str
    account_number: Optional[str] = ""

class Bank(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    account_number: Optional[str] = ""
    is_active: bool = True
    created_at: str
    updated_at: str

@api_router.get("/banks")
async def get_banks(user: dict = Depends(require_role(["admin"]))):
    """Get all banks"""
    banks = await db.banks.find({}, {"_id": 0}).to_list(None)
    # Sort by name
    banks.sort(key=lambda x: x['name'].lower())
    return banks

@api_router.post("/banks", response_model=Bank)
async def create_bank(bank: BankCreate, user: dict = Depends(require_role(["admin"]))):
    """Create a new bank"""
    # Check if bank name already exists
    existing = await db.banks.find_one({"name": {"$regex": f"^{bank.name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Bank with this name already exists")
    
    new_bank = {
        "id": str(uuid.uuid4()),
        "name": bank.name.strip(),
        "account_number": bank.account_number.strip() if bank.account_number else "",
        "is_active": True,
        "created_at": get_ist_now_iso(),
        "updated_at": get_ist_now_iso()
    }
    await db.banks.insert_one(new_bank)
    return new_bank

@api_router.put("/banks/{bank_id}", response_model=Bank)
async def update_bank(bank_id: str, bank: BankCreate, user: dict = Depends(require_role(["admin"]))):
    """Update a bank"""
    existing_bank = await db.banks.find_one({"id": bank_id})
    if not existing_bank:
        raise HTTPException(status_code=404, detail="Bank not found")
    
    # Check if new name conflicts with another bank
    name_conflict = await db.banks.find_one({
        "name": {"$regex": f"^{bank.name}$", "$options": "i"},
        "id": {"$ne": bank_id}
    })
    if name_conflict:
        raise HTTPException(status_code=400, detail="Bank with this name already exists")
    
    await db.banks.update_one(
        {"id": bank_id},
        {"$set": {
            "name": bank.name.strip(),
            "account_number": bank.account_number.strip() if bank.account_number else "",
            "updated_at": get_ist_now_iso()
        }}
    )
    updated_bank = await db.banks.find_one({"id": bank_id}, {"_id": 0})
    return updated_bank

@api_router.put("/banks/{bank_id}/status")
async def toggle_bank_status(bank_id: str, user: dict = Depends(require_role(["admin"]))):
    """Toggle bank active/inactive status"""
    bank = await db.banks.find_one({"id": bank_id})


# ==================== ATTENDANCE MODULE ====================
# Attendance - Get monthly attendance matrix

@api_router.get("/attendance")
async def get_attendance(year: int, month: int, user: dict = Depends(require_role(["admin", "hr"]))):
    """Get attendance matrix for a specific month - OPTIMIZED"""
    from calendar import monthrange
    
    # Get all active employees
    employees = await db.employees.find({"status": "active"}, {"_id": 0}).sort("name", 1).to_list(None)
    
    # Get number of days in the month
    num_days = monthrange(year, month)[1]
    dates = list(range(1, num_days + 1))
    
    # Check if this is a future month or current month
    today = datetime.now(IST)
    is_future_month = year > today.year or (year == today.year and month > today.month)
    is_current_month = year == today.year and month == today.month
    
    # OPTIMIZATION: Fetch all data at once
    # 1. Get all holidays for this month
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{num_days:02d}"
    holidays_cursor = db.holidays.find({
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0})
    holidays_list = await holidays_cursor.to_list(None)
    holiday_dates = {h["date"] for h in holidays_list}
    
    # 2. Get all approved leave applications for all employees
    leave_apps = await db.leave_applications.find({
        "status": "approved"
    }, {"_id": 0, "employee_id": 1, "leave_dates": 1}).to_list(None)
    
    # Build leave map: {employee_id: {date: leave_type}}
    leave_map = {}
    for leave_app in leave_apps:
        emp_id = leave_app.get("employee_id")
        if emp_id not in leave_map:
            leave_map[emp_id] = {}
        
        if leave_app.get("leave_dates"):
            for leave_date_entry in leave_app["leave_dates"]:
                date = leave_date_entry.get("date")
                leave_type = leave_date_entry.get("leave_type", "PL")
                if leave_type != "Rejected" and date >= start_date and date <= end_date:
                    leave_map[emp_id][date] = leave_type
    
    # 3. Get all work entries for this month (for OT calculation)
    work_entries = await db.work_entries.find({
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0, "employee_id": 1, "date": 1, "hours": 1}).to_list(None)
    
    # Build work hours map: {employee_id: {date: total_hours}}
    work_hours_map = {}
    for entry in work_entries:
        emp_id = entry.get("employee_id")
        date = entry.get("date")
        hours = entry.get("hours", 0)
        
        if emp_id not in work_hours_map:
            work_hours_map[emp_id] = {}
        if date not in work_hours_map[emp_id]:
            work_hours_map[emp_id][date] = 0
        
        work_hours_map[emp_id][date] += hours
    
    # Now build attendance matrix
    attendance = {}
    
    for emp in employees:
        emp_id = emp["employee_id"]
        attendance[emp_id] = {}
        
        # Get employee's joining date to check if they had joined yet
        joining_date_str = emp.get("joining_date", "")
        emp_joining_date = None
        if joining_date_str:
            try:
                emp_joining_date = datetime.fromisoformat(joining_date_str.replace('Z', '+00:00'))
                if emp_joining_date.tzinfo is not None:
                    emp_joining_date = emp_joining_date.replace(tzinfo=None)
            except:
                emp_joining_date = None
        
        for day in dates:
            date_str = f"{year}-{month:02d}-{day:02d}"
            current_date = datetime(year, month, day)
            day_of_week = current_date.weekday()
            
            is_holiday = date_str in holiday_dates
            is_weekend = day_of_week >= 5
            
            # Check if employee had joined by this date
            is_before_joining = False
            if emp_joining_date and current_date < emp_joining_date:
                is_before_joining = True
            
            # Check for future dates
            is_future_date = False
            if is_current_month:
                is_future_date = day > today.day
            elif is_future_month:
                is_future_date = True
            
            # If date is before employee's joining date, mark as "NJ" (Not Joined)
            if is_before_joining:
                attendance[emp_id][day] = "NJ"
            # Check if employee has leave on this date
            elif emp_id in leave_map and date_str in leave_map[emp_id]:
                attendance[emp_id][day] = leave_map[emp_id][date_str]
            elif (is_weekend or is_holiday) and not is_future_date:
                # Check if employee worked (OT)
                total_hours = work_hours_map.get(emp_id, {}).get(date_str, 0)
                
                if total_hours >= 8.5:
                    attendance[emp_id][day] = "OT"
                elif total_hours >= 4.5:
                    attendance[emp_id][day] = "OT/2"
                elif is_holiday:
                    attendance[emp_id][day] = "H"
                else:
                    attendance[emp_id][day] = "WO"
            elif is_weekend:
                attendance[emp_id][day] = "WO"
            elif is_holiday:
                attendance[emp_id][day] = "H"
            elif is_future_date:
                attendance[emp_id][day] = "-"
            else:
                attendance[emp_id][day] = "P"
    
    return {
        "year": year,
        "month": month,
        "num_days": num_days,
        "dates": dates,
        "employees": employees,
        "attendance": attendance,
        "sandwich_dates": await calculate_sandwich_dates(employees, attendance, year, month, num_days, holiday_dates)
    }

async def calculate_sandwich_dates(employees, attendance, year, month, num_days, holiday_dates):
    """Calculate sandwich leave dates for all employees"""
    sandwich_map = {}
    
    for emp in employees:
        emp_id = emp["employee_id"]
        emp_attendance = attendance.get(emp_id, {})
        
        # Build day types for sandwich detection
        day_types = []  # (day_num, date_str, status)
        for day in range(1, num_days + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            current_date = datetime(year, month, day)
            day_of_week = current_date.weekday()
            is_holiday = date_str in holiday_dates
            is_weekend = day_of_week >= 5
            
            status = emp_attendance.get(day, "P")
            
            # Classify the day
            if status == "NJ":
                day_types.append((day, date_str, 'notjoined'))
            elif status in ["PL", "CL", "PL/2", "CL/2", "PL/2 & CL/2", "Half PL", "Half CL"]:
                day_types.append((day, date_str, 'leave'))
            elif status in ["WO", "H"]:
                day_types.append((day, date_str, 'nonworking'))
            else:
                day_types.append((day, date_str, 'present'))
        
        # Detect sandwich leaves (same logic as salary calculation)
        # Group consecutive nonworking days (weekends/holidays)
        nw_groups = []
        i = 0
        while i < len(day_types):
            if day_types[i][2] == 'nonworking':
                start = i
                while i < len(day_types) and day_types[i][2] == 'nonworking':
                    i += 1
                nw_groups.append((start, i - 1))
            else:
                i += 1
        
        # Find sandwich patterns: leave-nonworking-leave
        sandwich_indices = set()
        for g in range(len(nw_groups) - 1):
            end_first = nw_groups[g][1]
            start_second = nw_groups[g + 1][0]
            between_start = end_first + 1
            between_end = start_second - 1
            if between_start > between_end:
                continue
            all_leave = all(day_types[j][2] == 'leave' for j in range(between_start, between_end + 1))
            if all_leave:
                for j in range(nw_groups[g][0], nw_groups[g][1] + 1):
                    sandwich_indices.add(j)
                for j in range(nw_groups[g + 1][0], nw_groups[g + 1][1] + 1):
                    sandwich_indices.add(j)
        
        # Get the actual day numbers that are sandwich
        sandwich_days = sorted([day_types[j][0] for j in sandwich_indices])
        if sandwich_days:
            sandwich_map[emp_id] = sandwich_days
    
    return sandwich_map

# Salary Calculation
# Late Coming - Admin marks employees who came late
@api_router.get("/late-coming")
async def get_late_coming(year: int, month: int, user: dict = Depends(require_role(["admin", "hr"]))):
    """Get late coming marks for all employees for a specific month"""
    from calendar import monthrange
    
    num_days = monthrange(year, month)[1]
    
    # Get all active employees
    employees = await db.employees.find({"status": "active"}, {"_id": 0, "employee_id": 1, "name": 1, "email": 1}).sort("name", 1).to_list(None)
    
    # Get late coming marks for this month
    month_key = f"{year}-{month:02d}"
    late_marks = await db.late_coming.find({"month_key": month_key}, {"_id": 0}).to_list(None)
    
    # Create a map: employee_id -> list of late dates
    late_map = {}
    for mark in late_marks:
        emp_id = mark.get("employee_id")
        if emp_id not in late_map:
            late_map[emp_id] = []
        late_map[emp_id].append(mark.get("day"))
    
    # Build response
    employee_data = []
    for emp in employees:
        emp_id = emp["employee_id"]
        employee_data.append({
            "employee_id": emp_id,
            "employee_name": emp.get("name", "Unknown"),
            "late_days": sorted(late_map.get(emp_id, []))
        })
    
    return {
        "year": year,
        "month": month,
        "num_days": num_days,
        "employees": employee_data,
        "total_late_marks": sum(len(e["late_days"]) for e in employee_data)
    }


@api_router.post("/late-coming")
async def toggle_late_coming(
    year: int,
    month: int,
    employee_id: str,
    day: int,
    user: dict = Depends(require_role(["admin"]))
):
    """Toggle late coming mark for an employee on a specific day"""
    from calendar import monthrange
    
    num_days = monthrange(year, month)[1]
    if day < 1 or day > num_days:
        raise HTTPException(status_code=400, detail=f"Invalid day: {day}")
    
    month_key = f"{year}-{month:02d}"
    
    # Check if mark exists
    existing = await db.late_coming.find_one({
        "month_key": month_key,
        "employee_id": employee_id,
        "day": day
    })
    
    if existing:
        # Remove the mark
        await db.late_coming.delete_one({
            "month_key": month_key,
            "employee_id": employee_id,
            "day": day
        })
        return {"action": "removed", "message": f"Late mark removed for day {day}"}
    else:
        # Add the mark
        await db.late_coming.insert_one({
            "month_key": month_key,
            "employee_id": employee_id,
            "day": day,
            "year": year,
            "month": month,
            "marked_by": user.get("username", "admin"),
            "marked_at": get_ist_now_iso()
        })
        return {"action": "added", "message": f"Late mark added for day {day}"}


@api_router.get("/late-coming/my")
async def get_my_late_coming(year: int, month: int, user: dict = Depends(require_role(["employee"]))):
    """Get employee's own late coming marks for a specific month"""
    # Get employee data
    employee = await db.employees.find_one({"email": user["email"]}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    emp_id = employee["employee_id"]
    month_key = f"{year}-{month:02d}"
    
    # Get late marks for this employee
    late_marks = await db.late_coming.find({
        "month_key": month_key,
        "employee_id": emp_id
    }, {"_id": 0, "day": 1}).to_list(None)
    
    late_days = sorted([m["day"] for m in late_marks])
    
    return {
        "year": year,
        "month": month,
        "late_days": late_days,
        "total_late": len(late_days)
    }


@api_router.get("/salary")
async def get_salary(year: int, month: int, user: dict = Depends(require_role(["admin", "hr"]))):
    """Calculate monthly salary for all active employees"""
    from calendar import monthrange

    num_days = monthrange(year, month)[1]
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{num_days:02d}"

    today = datetime.now(IST)
    is_future_month = year > today.year or (year == today.year and month > today.month)
    is_current_month = year == today.year and month == today.month

    employees = await db.employees.find({"status": "active"}, {"_id": 0}).sort("name", 1).to_list(None)

    # Load banks for mapping
    banks_list = await db.banks.find({}, {"_id": 0}).to_list(None)
    banks_map = {b["id"]: b["name"] for b in banks_list}

    holidays_list = await db.holidays.find({
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(None)
    holiday_dates = {h["date"] for h in holidays_list}

    leave_apps = await db.leave_applications.find({
        "status": "approved"
    }, {"_id": 0, "employee_id": 1, "leave_dates": 1}).to_list(None)

    leave_map = {}
    for leave_app in leave_apps:
        emp_id = leave_app.get("employee_id")
        if emp_id not in leave_map:
            leave_map[emp_id] = {}
        if leave_app.get("leave_dates"):
            for ld in leave_app["leave_dates"]:
                date = ld.get("date")
                leave_type = ld.get("leave_type", "PL")
                if leave_type != "Rejected" and date >= start_date and date <= end_date:
                    leave_map[emp_id][date] = leave_type

    work_entries = await db.work_entries.find({
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0, "employee_id": 1, "date": 1, "hours": 1, "is_compensation": 1}).to_list(None)

    work_hours_map = {}
    for entry in work_entries:
        # Skip compensation entries - they don't count as OT
        if entry.get("is_compensation", False):
            continue
        emp_id = entry.get("employee_id")
        date = entry.get("date")
        hours = entry.get("hours", 0)
        if emp_id not in work_hours_map:
            work_hours_map[emp_id] = {}
        if date not in work_hours_map[emp_id]:
            work_hours_map[emp_id][date] = 0
        work_hours_map[emp_id][date] += hours

    # Load salary adjustments for this month
    adjustments_list = await db.salary_adjustments.find({
        "year": year, "month": month
    }, {"_id": 0}).to_list(None)
    adjustments_map = {a["employee_id"]: a for a in adjustments_list}

    # Load late coming marks for this month
    month_key = f"{year}-{month:02d}"
    late_coming_list = await db.late_coming.find({"month_key": month_key}, {"_id": 0}).to_list(None)
    late_coming_map = {}
    for lc in late_coming_list:
        emp_id = lc.get("employee_id")
        if emp_id not in late_coming_map:
            late_coming_map[emp_id] = []
        late_coming_map[emp_id].append(lc.get("day"))

    salary_data = []
    for emp in employees:
        emp_id = emp["employee_id"]
        salary_str = emp.get("salary", "") or ""
        pt_str = emp.get("pt", "") or ""
        esic_str = emp.get("esic", "") or ""
        epf_str = emp.get("epf", "") or ""
        cpf_str = emp.get("cpf", "") or ""
        
        # Get employee's joining date for pro-rata calculation
        joining_date_str = emp.get("joining_date", "")
        emp_joining_date = None
        joining_day = 1  # Default to 1st of month if no joining date
        not_joined_days = 0  # Days before joining date
        
        if joining_date_str:
            try:
                emp_joining_date = datetime.fromisoformat(joining_date_str.replace('Z', '+00:00'))
                if emp_joining_date.tzinfo is not None:
                    emp_joining_date = emp_joining_date.replace(tzinfo=None)
                
                # Check if employee joined in this month
                if emp_joining_date.year == year and emp_joining_date.month == month:
                    joining_day = emp_joining_date.day
                    not_joined_days = joining_day - 1  # Days before joining
                elif emp_joining_date.year > year or (emp_joining_date.year == year and emp_joining_date.month > month):
                    # Employee hasn't joined yet in this month
                    not_joined_days = num_days
            except:
                pass

        try:
            salary = float(salary_str) if salary_str else 0
        except (ValueError, TypeError):
            salary = 0
        try:
            pt = float(pt_str) if pt_str else 0
        except (ValueError, TypeError):
            pt = 0
        try:
            esic = float(esic_str) if esic_str else 0
        except (ValueError, TypeError):
            esic = 0
        try:
            epf = float(epf_str) if epf_str else 0
        except (ValueError, TypeError):
            epf = 0
        try:
            cpf = float(cpf_str) if cpf_str else 0
        except (ValueError, TypeError):
            cpf = 0

        adj = adjustments_map.get(emp_id, {})
        other_income = float(adj.get("other_income", 0) or 0)
        extra_hours = float(adj.get("extra_hours", 0) or 0)

        cl_count = 0
        ot_count = 0

        # Build day classification for sandwich leave detection
        day_types = []  # (day_num, date_str, status)
        for day in range(1, num_days + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            current_date = datetime(year, month, day)
            day_of_week = current_date.weekday()
            is_holiday = date_str in holiday_dates
            is_weekend = day_of_week >= 5
            
            # Check if day is before employee's joining date
            is_before_joining = False
            if emp_joining_date and current_date < emp_joining_date:
                is_before_joining = True

            is_future_date = False
            if is_current_month:
                is_future_date = day > today.day
            elif is_future_month:
                is_future_date = True
            
            # Skip days before joining - mark as 'notjoined'
            if is_before_joining:
                day_types.append((day, date_str, 'notjoined'))
                continue

            # Check leave FIRST - approved leaves for future dates still count
            if emp_id in leave_map and date_str in leave_map[emp_id]:
                lt = leave_map[emp_id][date_str]
                if lt == "CL":
                    cl_count += 1
                elif lt == "Half CL":
                    cl_count += 0.5
                elif lt == "PL/2 & CL/2":
                    cl_count += 0.5
                day_types.append((day, date_str, 'leave'))
            elif is_weekend or is_holiday:
                if not is_future_date:
                    total_hours = work_hours_map.get(emp_id, {}).get(date_str, 0)
                    if total_hours >= 8.5:
                        ot_count += 1
                        day_types.append((day, date_str, 'present'))
                    elif total_hours >= 4.5:
                        ot_count += 0.5
                        day_types.append((day, date_str, 'present'))
                    else:
                        day_types.append((day, date_str, 'nonworking'))
                else:
                    # Future weekends/holidays still count as nonworking for sandwich detection
                    day_types.append((day, date_str, 'nonworking'))
            elif is_future_date:
                day_types.append((day, date_str, 'future'))
            else:
                day_types.append((day, date_str, 'present'))

        # Sandwich leave: find non-working groups and check if all
        # working days between adjacent groups are on leave
        nw_groups = []
        i = 0
        while i < len(day_types):
            if day_types[i][2] == 'nonworking':
                start = i
                while i < len(day_types) and day_types[i][2] == 'nonworking':
                    i += 1
                nw_groups.append((start, i - 1))
            else:
                i += 1

        sandwich_indices = set()
        for g in range(len(nw_groups) - 1):
            end_first = nw_groups[g][1]
            start_second = nw_groups[g + 1][0]
            between_start = end_first + 1
            between_end = start_second - 1
            if between_start > between_end:
                continue
            all_leave = all(day_types[j][2] == 'leave' for j in range(between_start, between_end + 1))
            if all_leave:
                for j in range(nw_groups[g][0], nw_groups[g][1] + 1):
                    sandwich_indices.add(j)
                for j in range(nw_groups[g + 1][0], nw_groups[g + 1][1] + 1):
                    sandwich_indices.add(j)

        sandwich_count = len(sandwich_indices)

        # Late coming calculation: first 2 are free, every 3 = 0.5 day deduction
        late_coming_days = late_coming_map.get(emp_id, [])
        late_coming_count = len(late_coming_days)
        # Deduction kicks in after 2 free late marks, then every 3 = 0.5 day
        late_coming_deduction_days = (late_coming_count // 3) * 0.5 if late_coming_count >= 3 else 0

        per_day = salary / num_days if num_days > 0 and salary > 0 else 0
        cl_amount = round(per_day * cl_count, 2)
        ot_amount = round(per_day * ot_count, 2)
        sandwich_amount = round(per_day * sandwich_count, 2)
        late_coming_amount = round(per_day * late_coming_deduction_days, 2)
        not_joined_amount = round(per_day * not_joined_days, 2)  # Deduction for days before joining
        per_hour = (per_day / 8.5) if per_day > 0 else 0
        extra_hours_amount = round(per_hour * extra_hours, 2)
        gross_salary = round(salary - pt - esic - epf - cpf - cl_amount - sandwich_amount - late_coming_amount - not_joined_amount + ot_amount + other_income + extra_hours_amount, 2)

        # Till Date Salary: what to pay if employee leaves today
        if is_current_month:
            future_days = num_days - today.day
        elif is_future_month:
            future_days = num_days
        else:
            future_days = 0
        future_amount = round(per_day * future_days, 2)
        td_salary = round(gross_salary - future_amount, 2)

        salary_data.append({
            "employee_id": emp_id,
            "employee_name": emp.get("name", "Unknown"),
            "joining_date": emp.get("joining_date", ""),
            "bank_id": emp.get("bank_id", ""),
            "bank_name": banks_map.get(emp.get("bank_id", ""), "No Bank"),
            "salary": salary,
            "pt": pt,
            "esic": esic,
            "epf": epf,
            "cpf": cpf,
            "cl_count": cl_count,
            "ot_count": ot_count,
            "cl_amount": cl_amount,
            "ot_amount": ot_amount,
            "sandwich_days": sandwich_count,
            "sandwich_amount": sandwich_amount,
            "late_coming_count": late_coming_count,
            "late_coming_deduction_days": late_coming_deduction_days,
            "late_coming_amount": late_coming_amount,
            "not_joined_days": not_joined_days,
            "not_joined_amount": not_joined_amount,
            "other_income": other_income,
            "extra_hours": extra_hours,
            "extra_hours_amount": extra_hours_amount,
            "gross_salary": gross_salary,
            "td_salary": td_salary,
            "future_days": future_days,
            "future_amount": future_amount,
            "num_days": num_days
        })

    return {
        "year": year,
        "month": month,
        "num_days": num_days,
        "salary_data": salary_data
    }

@api_router.put("/salary/adjustments")
async def save_salary_adjustment(
    data: dict = Body(...),
    user: dict = Depends(require_role(["admin"]))
):
    """Save other_income and extra_hours for an employee for a specific month"""
    emp_id = data.get("employee_id")
    year = data.get("year")
    month = data.get("month")
    other_income = data.get("other_income", 0)
    extra_hours = data.get("extra_hours", 0)

    if not emp_id or not year or not month:
        raise HTTPException(status_code=400, detail="employee_id, year, month required")

    await db.salary_adjustments.update_one(
        {"employee_id": emp_id, "year": year, "month": month},
        {"$set": {
            "employee_id": emp_id,
            "year": year,
            "month": month,
            "other_income": other_income,
            "extra_hours": extra_hours
        }},
        upsert=True
    )
    return {"message": "Adjustment saved"}



class SalarySheetRequest(BaseModel):
    bank_id: str
    year: int
    month: int
    sheet_name: str
    payment_date: str  # Format: DD-MMM-YYYY

@api_router.post("/salary/download-sheet")
async def download_salary_sheet(
    request: SalarySheetRequest,
    user: dict = Depends(require_role(["admin"]))
):
    """Download salary sheet in ICICI bank format for a specific bank"""
    
    # Get bank details
    bank = await db.banks.find_one({"id": request.bank_id}, {"_id": 0})
    if not bank:
        raise HTTPException(status_code=404, detail="Bank not found")
    
    debit_account = bank.get("account_number", "")
    if not debit_account:
        raise HTTPException(status_code=400, detail="Bank account number not set. Please update bank details.")
    
    # Get month name for remarks
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    month_name = month_names[request.month - 1]
    remarks = f"{month_name} Salary"
    
    # Get all active employees assigned to this bank with salary data
    employees = await db.employees.find(
        {"bank_id": request.bank_id, "status": "active"},
        {"_id": 0}
    ).to_list(1000)
    
    if not employees:
        raise HTTPException(status_code=404, detail="No active employees found for this bank")
    
    # Get salary data for these employees
    from calendar import monthrange
    num_days = monthrange(request.year, request.month)[1]
    start_date = f"{request.year}-{request.month:02d}-01"
    end_date = f"{request.year}-{request.month:02d}-{num_days:02d}"
    
    today = datetime.now(IST)
    is_future_month = request.year > today.year or (request.year == today.year and request.month > today.month)
    is_current_month = request.year == today.year and request.month == today.month
    
    # Get holidays
    holidays_list = await db.holidays.find({"date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}).to_list(None)
    holiday_dates = {h["date"] for h in holidays_list}
    
    # Get leave applications
    leave_apps = await db.leave_applications.find(
        {"status": "approved"},
        {"_id": 0, "employee_id": 1, "leave_dates": 1}
    ).to_list(None)
    
    leave_map = {}
    for la in leave_apps:
        emp_id = la["employee_id"]
        for ld in (la.get("leave_dates") or []):
            date = ld.get("date")
            lt = ld.get("leave_type", "PL")
            if lt != "Rejected" and date >= start_date and date <= end_date:
                if emp_id not in leave_map:
                    leave_map[emp_id] = {}
                leave_map[emp_id][date] = lt
    
    # Get work entries for OT calculation
    emp_ids = [e["employee_id"] for e in employees]
    work_entries = await db.work_entries.find(
        {"employee_id": {"$in": emp_ids}, "date": {"$gte": start_date, "$lte": end_date}},
        {"_id": 0, "employee_id": 1, "date": 1, "hours": 1, "is_compensation": 1}
    ).to_list(None)
    
    work_hours_map = {}
    for entry in work_entries:
        if entry.get("is_compensation", False):
            continue
        emp_id = entry.get("employee_id")
        d = entry.get("date")
        if emp_id not in work_hours_map:
            work_hours_map[emp_id] = {}
        if d not in work_hours_map[emp_id]:
            work_hours_map[emp_id][d] = 0
        work_hours_map[emp_id][d] += entry.get("hours", 0)
    
    # Get salary adjustments
    adjustments = await db.salary_adjustments.find(
        {"year": request.year, "month": request.month},
        {"_id": 0}
    ).to_list(None)
    adjustments_map = {a["employee_id"]: a for a in adjustments}
    
    # Get late coming marks
    month_key = f"{request.year}-{request.month:02d}"
    late_coming_list = await db.late_coming.find(
        {"month_key": month_key},
        {"_id": 0, "employee_id": 1, "day": 1}
    ).to_list(None)
    
    late_coming_map = {}
    for lc in late_coming_list:
        emp_id = lc["employee_id"]
        if emp_id not in late_coming_map:
            late_coming_map[emp_id] = []
        late_coming_map[emp_id].append(lc["day"])
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = request.sheet_name if request.sheet_name else "Sheet1"
    
    # Define headers (21 columns as per ICICI format)
    headers = [
        "Debit A/c Number",
        "Beneficiary A/c Number",
        "Beneficiary Name",
        "Amount",
        "Payment Type (Mandatory for all types of payments)",
        "Payment date",
        "IFSC Code",
        "Payable Location",
        "Print Location",
        "Beneficiary.Mobile No.",
        "Beneficiary email-id",
        "Bene Address 1",
        "Bene Address 2",
        "Bene Address 3",
        "Bene Address 4",
        "Add detail 1",
        "Add detail 2",
        "Add detail 3",
        "Add detail 4",
        "Add detail 5",
        "Remarks"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Calculate salary for each employee and write rows
    row_num = 2
    for emp in employees:
        emp_id = emp["employee_id"]
        
        # Get employee's joining date
        joining_date_str = emp.get("joining_date", "")
        emp_joining_date = None
        not_joined_days = 0
        
        if joining_date_str:
            try:
                emp_joining_date = datetime.fromisoformat(joining_date_str.replace('Z', '+00:00'))
                if emp_joining_date.tzinfo is not None:
                    emp_joining_date = emp_joining_date.replace(tzinfo=None)
                
                if emp_joining_date.year == request.year and emp_joining_date.month == request.month:
                    not_joined_days = emp_joining_date.day - 1
                elif emp_joining_date.year > request.year or (emp_joining_date.year == request.year and emp_joining_date.month > request.month):
                    not_joined_days = num_days
            except:
                pass
        
        # Get base values
        try:
            salary = float(emp.get("salary") or 0)
        except:
            salary = 0
        try:
            pt = float(emp.get("pt") or 0)
        except:
            pt = 0
        try:
            esic = float(emp.get("esic") or 0)
        except:
            esic = 0
        try:
            epf = float(emp.get("epf") or 0)
        except:
            epf = 0
        try:
            cpf = float(emp.get("cpf") or 0)
        except:
            cpf = 0
        
        adj = adjustments_map.get(emp_id, {})
        other_income = float(adj.get("other_income", 0) or 0)
        extra_hours = float(adj.get("extra_hours", 0) or 0)
        
        # Calculate CL, OT, and sandwich leaves
        cl_count = 0
        ot_count = 0
        day_types = []
        
        for day in range(1, num_days + 1):
            date_str = f"{request.year}-{request.month:02d}-{day:02d}"
            current_date = datetime(request.year, request.month, day)
            day_of_week = current_date.weekday()
            is_holiday = date_str in holiday_dates
            is_weekend = day_of_week >= 5
            
            is_before_joining = emp_joining_date and current_date < emp_joining_date
            is_future_date = (is_current_month and day > today.day) or is_future_month
            
            if is_before_joining:
                day_types.append('notjoined')
                continue
            
            if emp_id in leave_map and date_str in leave_map[emp_id]:
                lt = leave_map[emp_id][date_str]
                if lt == "CL":
                    cl_count += 1
                elif lt == "Half CL":
                    cl_count += 0.5
                elif lt == "PL/2 & CL/2":
                    cl_count += 0.5
                day_types.append('leave')
            elif is_weekend or is_holiday:
                if not is_future_date:
                    total_hours = work_hours_map.get(emp_id, {}).get(date_str, 0)
                    if total_hours >= 8.5:
                        ot_count += 1
                        day_types.append('present')
                    elif total_hours >= 4.5:
                        ot_count += 0.5
                        day_types.append('present')
                    else:
                        day_types.append('nonworking')
                else:
                    day_types.append('nonworking')
            elif is_future_date:
                day_types.append('future')
            else:
                day_types.append('present')
        
        # Calculate sandwich leaves
        nw_groups = []
        i = 0
        while i < len(day_types):
            if day_types[i] == 'nonworking':
                start = i
                while i < len(day_types) and day_types[i] == 'nonworking':
                    i += 1
                nw_groups.append((start, i - 1))
            else:
                i += 1
        
        sandwich_count = 0
        for g in range(len(nw_groups) - 1):
            end_first = nw_groups[g][1]
            start_second = nw_groups[g + 1][0]
            between_start = end_first + 1
            between_end = start_second - 1
            if between_start > between_end:
                continue
            all_leave = all(day_types[j] == 'leave' for j in range(between_start, between_end + 1))
            if all_leave:
                sandwich_count += (nw_groups[g][1] - nw_groups[g][0] + 1)
                sandwich_count += (nw_groups[g + 1][1] - nw_groups[g + 1][0] + 1)
        
        # Calculate late coming deduction
        late_coming_days = late_coming_map.get(emp_id, [])
        late_coming_count = len(late_coming_days)
        late_coming_deduction_days = (late_coming_count // 3) * 0.5 if late_coming_count >= 3 else 0
        
        # Calculate amounts
        per_day = salary / num_days if num_days > 0 and salary > 0 else 0
        cl_amount = round(per_day * cl_count, 2)
        ot_amount = round(per_day * ot_count, 2)
        sandwich_amount = round(per_day * sandwich_count, 2)
        late_coming_amount = round(per_day * late_coming_deduction_days, 2)
        not_joined_amount = round(per_day * not_joined_days, 2)
        per_hour = (per_day / 8.5) if per_day > 0 else 0
        extra_hours_amount = round(per_hour * extra_hours, 2)
        
        gross_salary = round(salary - pt - esic - epf - cpf - cl_amount - sandwich_amount - late_coming_amount - not_joined_amount + ot_amount + other_income + extra_hours_amount, 2)
        
        # Skip if gross salary is 0 or negative
        if gross_salary <= 0:
            continue
        
        # Write row
        ws.cell(row=row_num, column=1, value=debit_account)  # Debit A/c Number
        ws.cell(row=row_num, column=2, value=emp.get("bank_account_number", ""))  # Beneficiary A/c Number
        ws.cell(row=row_num, column=3, value=emp.get("name", ""))  # Beneficiary Name
        ws.cell(row=row_num, column=4, value=int(gross_salary))  # Amount (integer)
        ws.cell(row=row_num, column=5, value="I")  # Payment Type
        ws.cell(row=row_num, column=6, value=request.payment_date)  # Payment date
        ws.cell(row=row_num, column=21, value=remarks)  # Remarks
        
        row_num += 1
    
    if row_num == 2:
        raise HTTPException(status_code=404, detail="No employees with positive salary found")
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as downloadable file
    filename = f"{request.sheet_name or 'salary_sheet'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# Late Marks - Monthly late mark management with salary hold status
@api_router.get("/late-marks")
async def get_late_marks(year: int, month: int, user: dict = Depends(require_role(["admin", "hr"]))):
    """Get all employees with their late mark status for a specific month"""
    
    # Find all projects with status 'late'
    late_projects = await db.projects.find(
        {"status": "late"},
        {"_id": 0, "id": 1, "name": 1, "project_code": 1, "assigned_employees": 1, "client_username": 1, "start_date": 1, "end_date": 1}
    ).to_list(None)
    
    # Build map of employee_id -> list of late projects
    late_project_map = {}
    for project in late_projects:
        assigned_emps = project.get("assigned_employees", [])
        for emp_id in assigned_emps:
            if emp_id not in late_project_map:
                late_project_map[emp_id] = []
            late_project_map[emp_id].append({
                "project_id": project.get("id"),
                "project_name": project.get("name"),
                "project_code": project.get("project_code"),
                "client_name": project.get("client_username", ""),
                "start_date": project.get("start_date", ""),
                "end_date": project.get("end_date", "")
            })
    
    # Get all active employees
    employees = await db.employees.find(
        {"status": "active"},
        {"_id": 0, "employee_id": 1, "name": 1, "email": 1, "department_ids": 1}
    ).sort("name", 1).to_list(None)
    
    # Get departments for lookup
    departments = await db.departments.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(None)
    dept_map = {d["id"]: d["name"] for d in departments}
    
    # Get existing salary hold statuses for this month
    month_key = f"{year}-{month:02d}"
    salary_holds = await db.late_mark_salary_holds.find(
        {"month_key": month_key},
        {"_id": 0}
    ).to_list(None)
    salary_hold_map = {sh["employee_id"]: sh["salary_status"] for sh in salary_holds}
    
    # Build employee list
    employee_list = []
    for emp in employees:
        emp_id = emp["employee_id"]
        has_late_mark = emp_id in late_project_map
        late_projects_list = late_project_map.get(emp_id, [])
        
        # Get department names
        dept_names = [dept_map.get(d, "Unknown") for d in emp.get("department_ids", [])]
        
        # Determine salary status
        # If employee has late mark and no saved status, default to "hold"
        # If no late mark and no saved status, default to "active"
        if emp_id in salary_hold_map:
            salary_status = salary_hold_map[emp_id]
        else:
            salary_status = "hold" if has_late_mark else "active"
        
        employee_list.append({
            "employee_id": emp_id,
            "employee_name": emp.get("name", "Unknown"),
            "employee_email": emp.get("email", ""),
            "departments": dept_names,
            "has_late_mark": has_late_mark,
            "late_projects": late_projects_list,
            "late_project_count": len(late_projects_list),
            "salary_status": salary_status
        })
    
    # Sort: late mark employees first, then alphabetically
    employee_list.sort(key=lambda x: (not x["has_late_mark"], x["employee_name"].lower()))
    
    # Summary stats
    employees_with_late_marks = sum(1 for e in employee_list if e["has_late_mark"])
    employees_on_hold = sum(1 for e in employee_list if e["salary_status"] == "hold")
    
    return {
        "year": year,
        "month": month,
        "employees": employee_list,
        "total_employees": len(employee_list),
        "employees_with_late_marks": employees_with_late_marks,
        "employees_on_hold": employees_on_hold,
        "unique_late_projects": len(late_projects)
    }


# Update salary hold status for an employee
@api_router.put("/late-marks/salary-status")
async def update_salary_status(
    year: int,
    month: int,
    employee_id: str,
    salary_status: str,
    user: dict = Depends(require_role(["admin"]))
):
    """Update salary hold status for an employee for a specific month"""
    if salary_status not in ["hold", "active"]:
        raise HTTPException(status_code=400, detail="Status must be 'hold' or 'active'")
    
    month_key = f"{year}-{month:02d}"
    
    await db.late_mark_salary_holds.update_one(
        {"month_key": month_key, "employee_id": employee_id},
        {"$set": {
            "month_key": month_key,
            "employee_id": employee_id,
            "salary_status": salary_status,
            "updated_at": get_ist_now_iso(),
            "updated_by": user.get("username", "admin")
        }},
        upsert=True
    )
    
    return {"message": f"Salary status updated to {salary_status}"}


# Employee - Get my salary hold status
@api_router.get("/late-marks/my-status")
async def get_my_late_mark_status(year: int, month: int, user: dict = Depends(require_role(["employee"]))):
    """Get employee's own late mark and salary hold status for a specific month"""
    
    # Get employee data
    employee = await db.employees.find_one({"email": user["email"]}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    emp_id = employee["employee_id"]
    
    # Find all projects with status 'late' that this employee is assigned to
    late_projects = await db.projects.find(
        {"status": "late", "assigned_employees": emp_id},
        {"_id": 0, "id": 1, "name": 1, "project_code": 1, "client_username": 1}
    ).to_list(None)
    
    has_late_mark = len(late_projects) > 0
    
    # Get salary hold status for this month
    month_key = f"{year}-{month:02d}"
    salary_hold = await db.late_mark_salary_holds.find_one(
        {"month_key": month_key, "employee_id": emp_id},
        {"_id": 0}
    )
    
    # Determine salary status
    if salary_hold:
        salary_status = salary_hold.get("salary_status", "active")
    else:
        # Default: hold if has late mark, active otherwise
        salary_status = "hold" if has_late_mark else "active"
    
    return {
        "employee_id": emp_id,
        "employee_name": employee.get("name", ""),
        "year": year,
        "month": month,
        "has_late_mark": has_late_mark,
        "late_projects": [
            {
                "project_name": p.get("name"),
                "project_code": p.get("project_code"),
                "client_name": p.get("client_username", "")
            } for p in late_projects
        ],
        "salary_status": salary_status
    }


# Employee Attendance - View own attendance only
@api_router.get("/attendance/my")
async def get_my_attendance(year: int, month: int, user: dict = Depends(require_role(["employee"]))):
    """Get employee's own attendance for a specific month"""
    from calendar import monthrange
    
    # Get employee data
    employee = await db.employees.find_one({"email": user["email"]}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    emp_id = employee["employee_id"]
    
    # Get number of days in the month
    num_days = monthrange(year, month)[1]
    dates = list(range(1, num_days + 1))
    
    # Check if this is a future month or current month
    today = datetime.now(IST)
    is_future_month = year > today.year or (year == today.year and month > today.month)
    is_current_month = year == today.year and month == today.month
    
    # Get employee's joining date
    joining_date_str = employee.get("joining_date", "")
    emp_joining_date = None
    if joining_date_str:
        try:
            emp_joining_date = datetime.fromisoformat(joining_date_str.replace('Z', '+00:00'))
            if emp_joining_date.tzinfo is not None:
                emp_joining_date = emp_joining_date.replace(tzinfo=None)
        except:
            emp_joining_date = None
    
    # Get holidays
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{num_days:02d}"
    holidays_list = await db.holidays.find({"date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}).to_list(None)
    holiday_dates = {h["date"] for h in holidays_list}
    
    # Initialize attendance
    attendance = {}
    compensation_dates = []
    day_types = []  # For sandwich detection
    
    for day in dates:
        date_str = f"{year}-{month:02d}-{day:02d}"
        current_date = datetime(year, month, day)
        day_of_week = current_date.weekday()
        
        # Check if it's a holiday
        is_holiday = date_str in holiday_dates
        is_weekend = day_of_week >= 5
        
        # Check if day is before employee's joining date
        is_before_joining = False
        if emp_joining_date and current_date < emp_joining_date:
            is_before_joining = True
        
        # Check for future dates
        is_future_date = False
        if is_current_month:
            is_future_date = day > today.day
        elif is_future_month:
            is_future_date = True
        
        # If date is before employee's joining date, mark as "NJ" (Not Joined)
        if is_before_joining:
            attendance[day] = "NJ"
            day_types.append((day, date_str, 'notjoined'))
            continue
        
        # Check approved leave applications
        leave_found = False
        leave_apps = await db.leave_applications.find({
            "employee_id": emp_id,
            "status": "approved"
        }).to_list(None)
        
        for leave_app in leave_apps:
            if leave_app.get("leave_dates"):
                for leave_date_entry in leave_app["leave_dates"]:
                    if leave_date_entry.get("date") == date_str:
                        leave_type = leave_date_entry.get("leave_type", "PL")
                        if leave_type != "Rejected":
                            attendance[day] = leave_type
                            day_types.append((day, date_str, 'leave'))
                            leave_found = True
                            break
            if leave_found:
                break
        
        if not leave_found:
            if (is_weekend or is_holiday) and not is_future_date:
                work_entries = await db.work_entries.find({
                    "employee_id": emp_id,
                    "date": date_str
                }).to_list(None)
                
                total_hours = sum(entry.get("hours", 0) for entry in work_entries)
                has_compensation = any(entry.get("is_compensation", False) for entry in work_entries)
                
                if total_hours >= 8.5:
                    attendance[day] = "OT"
                    day_types.append((day, date_str, 'present'))
                elif total_hours >= 4.5:
                    attendance[day] = "OT/2"
                    day_types.append((day, date_str, 'present'))
                elif is_holiday:
                    attendance[day] = "H"
                    day_types.append((day, date_str, 'nonworking'))
                else:
                    attendance[day] = "WO"
                    day_types.append((day, date_str, 'nonworking'))
                
                if has_compensation:
                    compensation_dates.append(day)
            elif is_weekend:
                attendance[day] = "WO"
                day_types.append((day, date_str, 'nonworking'))
            elif is_holiday:
                attendance[day] = "H"
                day_types.append((day, date_str, 'nonworking'))
            elif is_future_date:
                attendance[day] = "-"
                day_types.append((day, date_str, 'future'))
            else:
                attendance[day] = "P"
                day_types.append((day, date_str, 'present'))
    
    # Detect sandwich leaves
    nw_groups = []
    i = 0
    while i < len(day_types):
        if day_types[i][2] == 'nonworking':
            start = i
            while i < len(day_types) and day_types[i][2] == 'nonworking':
                i += 1
            nw_groups.append((start, i - 1))
        else:
            i += 1
    
    sandwich_indices = set()
    for g in range(len(nw_groups) - 1):
        end_first = nw_groups[g][1]
        start_second = nw_groups[g + 1][0]
        between_start = end_first + 1
        between_end = start_second - 1
        if between_start > between_end:
            continue
        all_leave = all(day_types[j][2] == 'leave' for j in range(between_start, between_end + 1))
        if all_leave:
            for j in range(nw_groups[g][0], nw_groups[g][1] + 1):
                sandwich_indices.add(j)
            for j in range(nw_groups[g + 1][0], nw_groups[g + 1][1] + 1):
                sandwich_indices.add(j)
    
    sandwich_dates = sorted([day_types[j][0] for j in sandwich_indices])
    
    return {
        "year": year,
        "month": month,
        "num_days": num_days,
        "dates": dates,
        "employee": employee,
        "attendance": attendance,
        "compensation_dates": compensation_dates,
        "sandwich_dates": sandwich_dates
    }

@api_router.get("/salary/my")
async def get_my_salary(year: int, month: int, user: dict = Depends(require_role(["employee"]))):
    """Get employee's own salary calculation for a specific month"""
    from calendar import monthrange

    employee = await db.employees.find_one({"email": user["email"]}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    emp_id = employee["employee_id"]
    num_days = monthrange(year, month)[1]
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{num_days:02d}"

    today = datetime.now(IST)
    is_future_month = year > today.year or (year == today.year and month > today.month)
    is_current_month = year == today.year and month == today.month

    # Get employee's joining date for pro-rata calculation
    joining_date_str = employee.get("joining_date", "")
    emp_joining_date = None
    joining_day = 1  # Default to 1st of month if no joining date
    not_joined_days = 0  # Days before joining date
    
    if joining_date_str:
        try:
            emp_joining_date = datetime.fromisoformat(joining_date_str.replace('Z', '+00:00'))
            if emp_joining_date.tzinfo is not None:
                emp_joining_date = emp_joining_date.replace(tzinfo=None)
            
            # Check if employee joined in this month
            if emp_joining_date.year == year and emp_joining_date.month == month:
                joining_day = emp_joining_date.day
                not_joined_days = joining_day - 1  # Days before joining
            elif emp_joining_date.year > year or (emp_joining_date.year == year and emp_joining_date.month > month):
                # Employee hasn't joined yet in this month
                not_joined_days = num_days
        except:
            pass

    holidays_list = await db.holidays.find({"date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}).to_list(None)
    holiday_dates = {h["date"] for h in holidays_list}

    leave_apps = await db.leave_applications.find({"employee_id": emp_id, "status": "approved"}, {"_id": 0, "leave_dates": 1}).to_list(None)
    leave_map = {}
    for la in leave_apps:
        for ld in (la.get("leave_dates") or []):
            date = ld.get("date")
            lt = ld.get("leave_type", "PL")
            if lt != "Rejected" and date >= start_date and date <= end_date:
                leave_map[date] = lt

    work_entries = await db.work_entries.find({"employee_id": emp_id, "date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0, "date": 1, "hours": 1, "is_compensation": 1}).to_list(None)
    work_hours_map = {}
    for entry in work_entries:
        if entry.get("is_compensation", False):
            continue
        d = entry.get("date")
        if d not in work_hours_map:
            work_hours_map[d] = 0
        work_hours_map[d] += entry.get("hours", 0)

    try:
        salary = float(employee.get("salary") or 0)
    except (ValueError, TypeError):
        salary = 0
    try:
        pt = float(employee.get("pt") or 0)
    except (ValueError, TypeError):
        pt = 0
    try:
        esic = float(employee.get("esic") or 0)
    except (ValueError, TypeError):
        esic = 0
    try:
        epf = float(employee.get("epf") or 0)
    except (ValueError, TypeError):
        epf = 0
    try:
        cpf = float(employee.get("cpf") or 0)
    except (ValueError, TypeError):
        cpf = 0

    adj = await db.salary_adjustments.find_one({"employee_id": emp_id, "year": year, "month": month}, {"_id": 0})
    other_income = float((adj or {}).get("other_income", 0) or 0)
    extra_hours = float((adj or {}).get("extra_hours", 0) or 0)

    # Load late coming marks for this employee
    month_key = f"{year}-{month:02d}"
    late_coming_list = await db.late_coming.find({"month_key": month_key, "employee_id": emp_id}, {"_id": 0, "day": 1}).to_list(None)
    late_coming_days = sorted([lc["day"] for lc in late_coming_list])
    late_coming_count = len(late_coming_days)

    cl_count = 0
    ot_count = 0
    pl_count = 0
    cl_dates = []
    day_types = []

    for day in range(1, num_days + 1):
        date_str = f"{year}-{month:02d}-{day:02d}"
        current_date = datetime(year, month, day)
        day_of_week = current_date.weekday()
        is_holiday = date_str in holiday_dates
        is_weekend = day_of_week >= 5
        is_future_date = (is_current_month and day > today.day) or is_future_month
        
        # Check if day is before employee's joining date
        is_before_joining = False
        if emp_joining_date and current_date < emp_joining_date:
            is_before_joining = True
        
        # Skip days before joining - mark as 'notjoined'
        if is_before_joining:
            day_types.append((day, date_str, 'notjoined'))
            continue

        if date_str in leave_map:
            lt = leave_map[date_str]
            if lt == "CL":
                cl_count += 1
                cl_dates.append(day)
            elif lt == "Half CL":
                cl_count += 0.5
                cl_dates.append(day)
            elif lt == "PL/2 & CL/2":
                cl_count += 0.5
                cl_dates.append(day)
                pl_count += 0.5
            elif lt == "PL":
                pl_count += 1
            elif lt == "PL/2":
                pl_count += 0.5
            day_types.append((day, date_str, 'leave'))
        elif is_weekend or is_holiday:
            if not is_future_date:
                total_hours = work_hours_map.get(date_str, 0)
                if total_hours >= 8.5:
                    ot_count += 1
                    day_types.append((day, date_str, 'present'))
                elif total_hours >= 4.5:
                    ot_count += 0.5
                    day_types.append((day, date_str, 'present'))
                else:
                    day_types.append((day, date_str, 'nonworking'))
            else:
                day_types.append((day, date_str, 'nonworking'))
        elif is_future_date:
            day_types.append((day, date_str, 'future'))
        else:
            day_types.append((day, date_str, 'present'))

    # Sandwich leave detection
    nw_groups = []
    i = 0
    while i < len(day_types):
        if day_types[i][2] == 'nonworking':
            start = i
            while i < len(day_types) and day_types[i][2] == 'nonworking':
                i += 1
            nw_groups.append((start, i - 1))
        else:
            i += 1

    sandwich_indices = set()
    for g in range(len(nw_groups) - 1):
        end_first = nw_groups[g][1]
        start_second = nw_groups[g + 1][0]
        between_start = end_first + 1
        between_end = start_second - 1
        if between_start > between_end:
            continue
        if all(day_types[j][2] == 'leave' for j in range(between_start, between_end + 1)):
            for j in range(nw_groups[g][0], nw_groups[g][1] + 1):
                sandwich_indices.add(j)
            for j in range(nw_groups[g + 1][0], nw_groups[g + 1][1] + 1):
                sandwich_indices.add(j)

    sandwich_count = len(sandwich_indices)
    sandwich_dates = sorted([day_types[j][0] for j in sandwich_indices])

    # Late coming calculation: first 2 are free, every 3 = 0.5 day deduction
    late_coming_deduction_days = (late_coming_count // 3) * 0.5 if late_coming_count >= 3 else 0

    per_day = salary / num_days if num_days > 0 and salary > 0 else 0
    cl_amount = round(per_day * cl_count, 2)
    ot_amount = round(per_day * ot_count, 2)
    sandwich_amount = round(per_day * sandwich_count, 2)
    late_coming_amount = round(per_day * late_coming_deduction_days, 2)
    not_joined_amount = round(per_day * not_joined_days, 2)  # Deduction for days before joining
    per_hour = (per_day / 8.5) if per_day > 0 else 0
    extra_hours_amount = round(per_hour * extra_hours, 2)
    gross_salary = round(salary - pt - esic - epf - cpf - cl_amount - sandwich_amount - late_coming_amount - not_joined_amount + ot_amount + other_income + extra_hours_amount, 2)

    if is_current_month:
        future_days = num_days - today.day
    elif is_future_month:
        future_days = num_days
    else:
        future_days = 0
    future_amount = round(per_day * future_days, 2)
    td_salary = round(gross_salary - future_amount, 2)

    return {
        "salary": salary, "pt": pt, "esic": esic, "epf": epf, "cpf": cpf,
        "cl_count": cl_count, "pl_count": pl_count, "ot_count": ot_count,
        "cl_amount": cl_amount, "ot_amount": ot_amount,
        "cl_dates": sorted(cl_dates),
        "sandwich_days": sandwich_count, "sandwich_amount": sandwich_amount,
        "sandwich_dates": sandwich_dates,
        "late_coming_count": late_coming_count,
        "late_coming_days": late_coming_days,
        "late_coming_deduction_days": late_coming_deduction_days,
        "late_coming_amount": late_coming_amount,
        "not_joined_days": not_joined_days,
        "not_joined_amount": not_joined_amount,
        "other_income": other_income, "extra_hours": extra_hours,
        "extra_hours_amount": extra_hours_amount,
        "gross_salary": gross_salary,
        "td_salary": td_salary, "future_days": future_days, "future_amount": future_amount,
        "num_days": num_days
    }



logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

@app.on_event("startup")
async def startup_db():
    # Create default admin user if not exists
    admin_exists = await db.users.find_one({"username": "renish"})
    if not admin_exists:
        admin_user = {
            "id": str(uuid.uuid4()),
            "username": "renish",
            "email": "renish@zestbrains.com",
            "password_hash": hash_password("Zb@0075588"),
            "role": "admin",
            "employee_id": "ADMIN001",
            "is_active": True,
            "created_at": get_ist_now_iso()
        }
        await db.users.insert_one(admin_user)
        logger.info("Default admin user created: username=renish, password=Zb@0075588")

# Include router at the END after all routes are defined
app.include_router(api_router)
